"""
ESGIntel — AI-Assisted ESG Due Diligence Platform
Streamlit frontend — calls FastAPI backend at localhost:8000
Run: streamlit run frontend/app.py
"""

import time
import requests
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd

# ──────────────────────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ESGIntel — AI-Assisted ESG Due Diligence",
    page_icon="🌱",
    layout="wide",
    initial_sidebar_state="expanded",
)

API_BASE = "http://localhost:8000"

# ──────────────────────────────────────────────────────────────────────────────
# Custom CSS — Bloomberg/MSCI aesthetic
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Tokens ── */
:root {
  --accent:       #3154ff;
  --accent-light: #e8f0ff;
  --red:          #e24141;
  --red-light:    #fff2f2;
  --amber:        #d17f1e;
  --amber-light:  #fff4de;
  --green:        #2f855a;
  --green-light:  #e9f7ee;
  --blue:         #2563eb;
  --blue-light:   #e7f0ff;
  --text-primary: #102a43;
  --text-muted:   #627d98;
  --border:       rgba(145,158,171,0.18);
  --surface:      #ffffff;
  --surface-2:    #f6f8ff;
  --bg:           #f4f7fb;
  --radius:       14px;
}

/* Main bg */
.stApp { background: #f4f7fb; }

/* Hide default Streamlit header/footer */
#MainMenu, footer, header { visibility: hidden; }

/* Sidebar styling */
[data-testid="stSidebar"] {
  background: rgba(255,255,255,0.97);
  border-right: 1px solid rgba(145,158,171,0.14);
}
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
  font-size: 13px;
  color: #627d98;
}

/* Metric cards */
[data-testid="stMetric"] {
  background: white;
  border: 1px solid rgba(145,158,171,0.14);
  border-radius: 14px;
  padding: 16px 18px;
  box-shadow: 0 4px 16px rgba(15,25,35,0.06);
}
[data-testid="stMetricLabel"] { font-size: 12px !important; color: #627d98 !important; font-weight: 600 !important; text-transform: uppercase; letter-spacing: 0.5px; }
[data-testid="stMetricValue"] { font-size: 28px !important; font-weight: 800 !important; color: #102a43 !important; }

/* Expander */
[data-testid="stExpander"] {
  background: white;
  border: 1px solid rgba(145,158,171,0.14);
  border-radius: 10px;
  margin-bottom: 6px;
}

/* Tabs */
[data-testid="stTabs"] [role="tab"] {
  font-size: 13px;
  font-weight: 600;
  padding: 8px 16px;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
  color: #3154ff;
  border-bottom: 2px solid #3154ff;
}

/* Buttons */
.stButton > button {
  border-radius: 999px;
  font-weight: 700;
  font-size: 14px;
  transition: all 0.15s;
}

/* DataFrames */
[data-testid="stDataFrame"] {
  border-radius: 10px;
  overflow: hidden;
}

/* Form inputs */
[data-testid="stTextInput"] input,
[data-testid="stSelectbox"] select,
[data-testid="stTextArea"] textarea {
  border-radius: 8px;
  border: 1px solid rgba(145,158,171,0.3);
  font-size: 14px;
}
</style>

<style>
/* ── Utility classes ── */
.badge-red   { display:inline-block; padding:3px 10px; border-radius:999px; font-size:11px; font-weight:700; background:#fff2f2; color:#e24141; }
.badge-amber { display:inline-block; padding:3px 10px; border-radius:999px; font-size:11px; font-weight:700; background:#fff4de; color:#d17f1e; }
.badge-green { display:inline-block; padding:3px 10px; border-radius:999px; font-size:11px; font-weight:700; background:#e9f7ee; color:#2f855a; }
.badge-blue  { display:inline-block; padding:3px 10px; border-radius:999px; font-size:11px; font-weight:700; background:#e7f0ff; color:#2563eb; }
.badge-gray  { display:inline-block; padding:3px 10px; border-radius:999px; font-size:11px; font-weight:700; background:#f1f5f9; color:#627d98; }

.card-box {
  background: white;
  border: 1px solid rgba(145,158,171,0.14);
  border-radius: 14px;
  padding: 16px 20px;
  box-shadow: 0 4px 16px rgba(15,25,35,0.06);
  margin-bottom: 12px;
}

.kpi-card {
  background: white;
  border-radius: 14px;
  padding: 18px 20px;
  border: 1px solid rgba(145,158,171,0.14);
  box-shadow: 0 4px 16px rgba(15,25,35,0.06);
  text-align: center;
}
.kpi-label { font-size: 11px; font-weight: 700; color: #627d98; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 6px; }
.kpi-value { font-size: 32px; font-weight: 800; color: #102a43; line-height: 1; }
.kpi-sub   { font-size: 12px; color: #627d98; margin-top: 4px; }

.signal-positive { border-left: 3px solid #2f855a; background: #e9f7ee; padding: 8px 12px; border-radius: 0 8px 8px 0; margin-bottom: 6px; }
.signal-negative { border-left: 3px solid #e24141; background: #fff2f2; padding: 8px 12px; border-radius: 0 8px 8px 0; margin-bottom: 6px; }
.signal-neutral  { border-left: 3px solid #2563eb; background: #e7f0ff; padding: 8px 12px; border-radius: 0 8px 8px 0; margin-bottom: 6px; }
.signal-label { font-size: 13px; font-weight: 700; color: #102a43; }
.signal-desc  { font-size: 12px; color: #627d98; margin-top: 2px; }

.evidence-box { background: #f6f8ff; border: 1px solid rgba(145,158,171,0.2); border-radius: 8px; padding: 10px 12px; margin-top: 8px; font-size: 12px; }
.evidence-source { font-weight: 700; color: #2563eb; margin-bottom: 3px; }
.evidence-text { color: #334e68; line-height: 1.5; }
.evidence-conf { color: #627d98; margin-top: 3px; font-style: italic; }

.risk-row { background: white; border: 1px solid rgba(145,158,171,0.14); border-radius: 10px; padding: 12px 16px; margin-bottom: 6px; cursor: pointer; }
.risk-name { font-size: 14px; font-weight: 700; color: #102a43; }
.risk-meta { font-size: 12px; color: #627d98; margin-top: 2px; }

.progress-step { display: flex; align-items: center; gap: 10px; padding: 6px 0; font-size: 13px; color: #334e68; }
.progress-done { color: #2f855a; font-weight: 600; }

.page-title { font-size: 26px; font-weight: 800; color: #102a43; letter-spacing: -0.4px; margin-bottom: 4px; }
.page-sub { font-size: 14px; color: #627d98; margin-bottom: 20px; }

/* ── Topbar ── */
.topbar-wrap {
  display:flex; align-items:center; gap:14px;
  background:#ffffff; border:1px solid rgba(145,158,171,0.14);
  border-radius:14px; padding:10px 18px; margin-bottom:14px;
  box-shadow:0 4px 16px rgba(15,25,35,0.05);
}
.topbar-brand { display:flex; align-items:center; gap:10px; }
.topbar-brand .logo { width:30px; height:30px; background:#3154ff; border-radius:8px;
  display:flex; align-items:center; justify-content:center; font-size:15px; }
.topbar-brand .name { font-weight:800; font-size:16px; color:#102a43; letter-spacing:-.3px; }
.company-chip { display:inline-block; padding:4px 12px; border-radius:999px; font-size:12.5px;
  font-weight:800; background:#eef2ff; color:#3154ff; }
.nace-chip { display:inline-block; padding:4px 12px; border-radius:999px; font-size:12px;
  font-weight:600; background:#f1f5f9; color:#627d98; }
.topbar-meta { font-size:11.5px; color:#94a3b8; }
.conf-badge { display:inline-flex; align-items:center; gap:6px; padding:4px 12px; border-radius:999px;
  font-size:12px; font-weight:700; background:#f6f8ff; border:1px solid rgba(145,158,171,0.18); }
.conf-badge .k { color:#627d98; text-transform:uppercase; letter-spacing:.4px; font-size:10.5px; }

/* ── Auth screen ── */
.auth-hero { text-align:center; margin: 18px auto 4px; }
.auth-hero .logo { width:52px; height:52px; background:#3154ff; border-radius:14px;
  display:inline-flex; align-items:center; justify-content:center; font-size:26px; margin-bottom:10px;
  box-shadow:0 8px 24px rgba(49,84,255,0.28); }
.auth-hero .name { font-size:26px; font-weight:800; color:#102a43; letter-spacing:-.5px; }
.auth-hero .sub { font-size:13px; color:#627d98; margin-top:2px; }
.auth-badge-row { display:flex; gap:8px; justify-content:center; flex-wrap:wrap; margin-top:14px; }
.auth-badge { font-size:11px; color:#627d98; background:#f1f5f9; border-radius:999px; padding:4px 10px; font-weight:600; }
</style>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# Session state initialisation
# ──────────────────────────────────────────────────────────────────────────────
def _init_state():
    defaults = {
        "page":        "input",
        "analysis_id": None,
        "result":      None,
        "company_name": "",
        "nace_code":   "C24",
        "website_url": "",
        "country":     "Italy",
        "employees":   "1,000 – 4,999",
        "revenue":     "€1B – €5B",
        "step":        1,
        # ── Auth ──
        "auth_token":  None,     # set after login/signup/verify; gates the whole app
        "user":        None,     # cached GET /api/auth/me payload
        "auth_mode":   "signin", # signin | signup | code
        "pending_email": "",     # email awaiting verification code
        "dev_code":    "",       # dev-mode verification code echoed by backend
        # ── Data validation (wizard step 3) ──
        "validation_rows": None, # cached /api/validate result
        "validation_done": False,
        # ── Materiality: session-only score overrides keyed by topic name ──
        "materiality_overrides": {},
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_state()


# ──────────────────────────────────────────────────────────────────────────────
# API client
# ──────────────────────────────────────────────────────────────────────────────
def api_submit(company_name, website_url, nace_code, country, employees, revenue, files):
    data = {
        "company_name": company_name,
        "website_url":  website_url or "",
        "nace_code":    nace_code,
        "country":      country,
        "employees":    employees,
        "revenue":      revenue,
    }
    file_list = []
    if files:
        for f in files:
            try:
                f.seek(0)  # validate step may have already read the buffer
            except Exception:
                pass
            file_list.append(("documents", (f.name, f.read(), f.type or "application/octet-stream")))
    else:
        # Send a stub file so the endpoint accepts it
        stub = b"Please analyse ESG risk based on publicly available information only."
        file_list.append(("documents", ("stub.txt", stub, "text/plain")))

    resp = requests.post(f"{API_BASE}/api/analyze", data=data, files=file_list, timeout=30)
    resp.raise_for_status()
    return resp.json()

def api_status(analysis_id):
    resp = requests.get(f"{API_BASE}/api/analyses/{analysis_id}/status", timeout=10)
    resp.raise_for_status()
    return resp.json()

def api_result(analysis_id):
    resp = requests.get(f"{API_BASE}/api/analyses/{analysis_id}", timeout=10)
    resp.raise_for_status()
    return resp.json()

def api_list():
    resp = requests.get(f"{API_BASE}/api/analyses", timeout=10)
    resp.raise_for_status()
    return resp.json()

def api_delete(analysis_id):
    resp = requests.delete(f"{API_BASE}/api/analyses/{analysis_id}", timeout=10)
    resp.raise_for_status()
    return resp.json()

def backend_ok():
    try:
        r = requests.get(f"{API_BASE}/", timeout=3)
        return r.status_code == 200
    except Exception:
        return False


# ── Auth API ──────────────────────────────────────────────────────────────────
def _auth_error(resp):
    """Pull a human-readable message out of a FastAPI error response."""
    try:
        body = resp.json()
        return body.get("detail") or body.get("message") or resp.text
    except Exception:
        return resp.text or f"HTTP {resp.status_code}"

def api_login(email, password):
    resp = requests.post(f"{API_BASE}/api/auth/login",
                         json={"email": email, "password": password}, timeout=15)
    if resp.status_code >= 400:
        raise RuntimeError(_auth_error(resp))
    return resp.json()

def api_signup(first_name, last_name, email, password, company, role):
    resp = requests.post(f"{API_BASE}/api/auth/signup", json={
        "firstName": first_name, "lastName": last_name, "email": email,
        "password": password, "company": company, "role": role,
    }, timeout=15)
    if resp.status_code >= 400:
        raise RuntimeError(_auth_error(resp))
    return resp.json()

def api_send_code(email, purpose="verify"):
    resp = requests.post(f"{API_BASE}/api/auth/send-code",
                         json={"email": email, "purpose": purpose}, timeout=15)
    if resp.status_code >= 400:
        raise RuntimeError(_auth_error(resp))
    return resp.json()

def api_verify_code(email, code, purpose="verify"):
    resp = requests.post(f"{API_BASE}/api/auth/verify-code",
                         json={"email": email, "code": code, "purpose": purpose}, timeout=15)
    if resp.status_code >= 400:
        raise RuntimeError(_auth_error(resp))
    return resp.json()

def api_me(token):
    resp = requests.get(f"{API_BASE}/api/auth/me",
                        headers={"Authorization": f"Bearer {token}"}, timeout=10)
    if resp.status_code >= 400:
        raise RuntimeError(_auth_error(resp))
    return resp.json()

def api_logout(token):
    try:
        requests.post(f"{API_BASE}/api/auth/logout",
                      headers={"Authorization": f"Bearer {token}"}, timeout=8)
    except Exception:
        pass  # best-effort — we clear local session regardless

def _finish_login(data):
    """Given a login/verify response ({token, user}), cache it in session_state."""
    token = data.get("token") or data.get("access_token")
    if not token:
        raise RuntimeError("No token returned by the server.")
    st.session_state.auth_token = token
    user = data.get("user")
    if not user:
        try:
            user = api_me(token)
        except Exception:
            user = {}
    st.session_state.user = user or {}


# ── Data validation API (wizard step 3) ───────────────────────────────────────
def api_validate(company_name, website_url, files):
    """Pre-run data extraction/validation. Returns {rows:[...]}."""
    data = {
        "company_name": company_name,
        "website_url":  website_url or "",
    }
    file_list = []
    if files:
        for f in files:
            try:
                f.seek(0)
            except Exception:
                pass
            file_list.append(("documents", (f.name, f.read(), f.type or "application/octet-stream")))
    resp = requests.post(f"{API_BASE}/api/validate", data=data,
                         files=file_list or None, timeout=120)
    resp.raise_for_status()
    return resp.json()


# ── Documents / Engagement / Canva API ─────────────────────────────────────────
DOC_MIME_BY_EXT = {
    "pdf":  "application/pdf",
    "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
# extension → (Document-Library category, icon, badge css class, accent colour)
DOC_CATEGORY_BY_EXT = {
    "pdf":  ("One-Pagers",     "📄", "badge-blue",  "#3154ff"),
    "pptx": ("Presentations",  "📊", "badge-blue",  "#7c3aed"),
    "docx": ("Policies",       "📋", "badge-red",   "#e24141"),
    "xlsx": ("Data Templates", "📗", "badge-green", "#2f855a"),
}
DOC_CATEGORY_ORDER = ["One-Pagers", "Presentations", "Policies", "Data Templates"]

def api_manifest():
    resp = requests.get(f"{API_BASE}/api/documents/manifest", timeout=10)
    resp.raise_for_status()
    return resp.json()

def api_engagement_file(fname):
    resp = requests.get(f"{API_BASE}/api/documents/engagement",
                        params={"file": fname}, timeout=90)
    resp.raise_for_status()
    return resp.content

def api_report_doc(rtype, fmt):
    """GET /api/documents/report?type=&fmt= — returns docx/pdf bytes."""
    resp = requests.get(f"{API_BASE}/api/documents/report",
                        params={"type": rtype, "fmt": fmt}, timeout=90)
    resp.raise_for_status()
    return resp.content

def api_canva_request(deck_type):
    resp = requests.post(f"{API_BASE}/api/documents/canva/request",
                         params={"type": deck_type}, timeout=20)
    resp.raise_for_status()
    return resp.json()

def api_canva_status(job_id):
    resp = requests.get(f"{API_BASE}/api/documents/canva/status/{job_id}", timeout=10)
    resp.raise_for_status()
    return resp.json()

def api_canva_latest(deck_type):
    resp = requests.get(f"{API_BASE}/api/documents/canva/latest",
                        params={"type": deck_type}, timeout=10)
    resp.raise_for_status()
    return resp.json()


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
NACE_OPTIONS = [
    "A01 — Crop and Animal Production", "A02 — Forestry and Logging",
    "B05 — Mining of Hard Coal", "B06 — Extraction of Crude Petroleum and Natural Gas",
    "C10 — Manufacture of Food Products", "C13 — Manufacture of Textiles",
    "C20 — Manufacture of Chemicals and Chemical Products",
    "C21 — Manufacture of Pharmaceutical Products",
    "C22 — Manufacture of Rubber and Plastic Products",
    "C24 — Manufacture of Basic Metals",
    "C25 — Manufacture of Fabricated Metal Products",
    "C26 — Manufacture of Computer and Electronic Products",
    "C27 — Manufacture of Electrical Equipment",
    "C28 — Manufacture of Machinery and Equipment",
    "C29 — Manufacture of Motor Vehicles",
    "D35 — Electricity, Gas, Steam and Air Conditioning Supply",
    "E36 — Water Collection, Treatment and Supply",
    "E38 — Waste Collection, Treatment and Disposal",
    "F41 — Construction of Buildings", "F42 — Civil Engineering",
    "G46 — Wholesale Trade", "G47 — Retail Trade",
    "H49 — Land Transport and Transport via Pipelines",
    "H50 — Water Transport", "H51 — Air Transport",
    "I55 — Accommodation", "I56 — Food and Beverage Service Activities",
    "J62 — Computer Programming and Consultancy",
    "J63 — Information Service Activities",
    "K64 — Financial Service Activities",
    "K65 — Insurance and Pension Funding",
    "L68 — Real Estate Activities",
    "M69 — Legal and Accounting Activities",
    "M70 — Management Consultancy",
    "M71 — Architectural and Engineering Activities",
    "M72 — Scientific Research and Development",
    "N77 — Rental and Leasing Activities",
    "P85 — Education", "Q86 — Human Health Activities",
]

COUNTRIES = [
    "Austria", "Belgium", "Brazil", "Canada", "China", "Denmark", "Finland",
    "France", "Germany", "India", "Indonesia", "Italy", "Japan", "Mexico",
    "Netherlands", "Norway", "Poland", "Portugal", "South Africa",
    "South Korea", "Spain", "Sweden", "Switzerland", "Turkey",
    "United Kingdom", "United States", "Other",
]

EMPLOYEE_RANGES = ["1 – 49", "50 – 249", "250 – 999", "1,000 – 4,999",
                   "5,000 – 9,999", "10,000 – 49,999", "50,000+"]

REVENUE_RANGES  = ["Under €10M", "€10M – €50M", "€50M – €250M",
                   "€250M – €1B", "€1B – €5B", "€5B – €20B", "Over €20B"]

def score_color(score):
    """Return hex color for a 0-100 risk score."""
    if score is None:
        return "#627d98"
    if score >= 70:
        return "#e24141"
    if score >= 40:
        return "#d17f1e"
    return "#2f855a"

def score_badge_class(score):
    if score is None:
        return "badge-gray"
    if score >= 70:
        return "badge-red"
    if score >= 40:
        return "badge-amber"
    return "badge-green"

def severity_badge(sev):
    m = {"Critical": "badge-red", "High": "badge-red",
         "Medium": "badge-amber", "Low": "badge-green"}
    cls = m.get(sev, "badge-gray")
    return f'<span class="{cls}">{sev}</span>'

def confidence_badge(conf):
    m = {"high": "badge-green", "medium": "badge-amber", "low": "badge-red"}
    cls = m.get((conf or "").lower(), "badge-gray")
    return f'<span class="{cls}">{(conf or "").title()}</span>'

def gauge_chart(value, title, max_val=100):
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=value,
        title={"text": title, "font": {"size": 13, "color": "#627d98"}},
        number={"font": {"size": 28, "color": score_color(value), "family": "Inter"}},
        gauge={
            "axis": {"range": [0, max_val], "tickwidth": 1},
            "bar": {"color": score_color(value)},
            "bgcolor": "white",
            "borderwidth": 0,
            "steps": [
                {"range": [0, 40],  "color": "#e9f7ee"},
                {"range": [40, 70], "color": "#fff4de"},
                {"range": [70, 100],"color": "#fff2f2"},
            ],
        },
    ))
    fig.update_layout(
        height=180, margin=dict(l=20, r=20, t=30, b=0),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif"),
    )
    return fig

def radar_chart(data: dict):
    cats  = list(data.keys())
    vals  = list(data.values())
    vals += [vals[0]]
    cats += [cats[0]]
    fig = go.Figure(go.Scatterpolar(
        r=vals, theta=cats, fill="toself",
        fillcolor="rgba(49,84,255,0.12)", line=dict(color="#3154ff", width=2),
    ))
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
        showlegend=False,
        height=300,
        margin=dict(l=20, r=20, t=20, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
    )
    return fig

def materiality_matrix(topics: list):
    if not topics:
        return None
    df = pd.DataFrame(topics)
    df["financial_score"] = pd.to_numeric(df.get("financial_score"), errors="coerce").fillna(3)
    df["impact_score"]    = pd.to_numeric(df.get("impact_score"),    errors="coerce").fillna(3)
    color_map = {"environmental": "#2f855a", "social": "#2563eb", "governance": "#d17f1e"}
    df["color"] = df.get("category", "").map(color_map).fillna("#627d98")

    fig = go.Figure()
    for cat, grp in df.groupby("category"):
        fig.add_trace(go.Scatter(
            x=grp["financial_score"], y=grp["impact_score"],
            mode="markers+text",
            marker=dict(size=16, color=color_map.get(cat, "#627d98"), opacity=0.85,
                        line=dict(width=1, color="white")),
            text=grp["topic"], textposition="top center",
            textfont=dict(size=10, color="#102a43"),
            name=cat.title(),
        ))

    fig.add_shape(type="line", x0=3, y0=0, x1=3, y1=5.5,
                  line=dict(dash="dot", color="#cbd5e0", width=1))
    fig.add_shape(type="line", x0=0, y0=3, x1=5.5, y1=3,
                  line=dict(dash="dot", color="#cbd5e0", width=1))

    fig.update_layout(
        xaxis=dict(title="Financial Materiality →", range=[0, 5.5], tickvals=[1,2,3,4,5]),
        yaxis=dict(title="Impact Materiality →",    range=[0, 5.5], tickvals=[1,2,3,4,5]),
        height=440,
        margin=dict(l=40, r=20, t=20, b=40),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#f6f8ff",
        legend=dict(orientation="h", y=-0.15),
        font=dict(family="Inter, sans-serif", size=12),
    )
    return fig

def risk_bar_chart(risks: list):
    if not risks:
        return None
    df = pd.DataFrame(risks)[["name", "score", "category", "severity"]].copy()
    df["score"] = pd.to_numeric(df["score"], errors="coerce").fillna(0)
    df = df.sort_values("score", ascending=True).tail(12)
    color_map = {"environmental": "#2f855a", "social": "#2563eb",
                 "governance": "#d17f1e", "climate": "#6366f1"}
    df["color"] = df["category"].map(color_map).fillna("#627d98")

    fig = go.Figure(go.Bar(
        x=df["score"], y=df["name"],
        orientation="h",
        marker_color=df["color"],
        text=df["score"].astype(int).astype(str),
        textposition="outside",
    ))
    fig.update_layout(
        height=max(300, len(df) * 36),
        xaxis=dict(range=[0, 105], title="Risk Score (0–100)"),
        yaxis=dict(tickfont=dict(size=12)),
        margin=dict(l=0, r=40, t=10, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif", size=12),
        showlegend=False,
    )
    return fig


ROLE_OPTIONS = [
    "Select role…",
    "Institutional Investor / Fund Manager",
    "ESG / Research Analyst",
    "ESG Consultant / Advisor",
    "Corporate Sustainability Officer",
    "Credit / Fixed Income Analyst",
    "Other",
]


# ──────────────────────────────────────────────────────────────────────────────
# Auth screen (shown when no auth_token is present — gates the whole app)
# ──────────────────────────────────────────────────────────────────────────────
def render_auth():
    # Hide the sidebar entirely on the auth screen
    st.markdown("<style>[data-testid='stSidebar']{display:none;}</style>", unsafe_allow_html=True)

    _, mid, _ = st.columns([1, 1.4, 1])
    with mid:
        st.markdown("""
        <div class="auth-hero">
          <div class="logo">🌱</div>
          <div class="name">ESGIntel</div>
          <div class="sub">AI-Assisted ESG Due Diligence Platform</div>
        </div>
        """, unsafe_allow_html=True)

        with st.container(border=True):
            if not backend_ok():
                st.warning("⚠️ Backend offline — start FastAPI at :8000 to sign in. "
                           "You can still use developer access below.")

            tab_in, tab_up, tab_code = st.tabs(["Sign In", "Create Account", "Email Code"])

            # ── Sign In ──
            with tab_in:
                st.markdown("##### Welcome back")
                st.caption("Enter your email and password to continue")
                with st.form("login_form"):
                    email = st.text_input("Email address", placeholder="you@firm.com")
                    pw    = st.text_input("Password", type="password", placeholder="••••••••")
                    submit = st.form_submit_button("Sign In →", type="primary", use_container_width=True)
                if submit:
                    if not email or not pw:
                        st.error("Enter your email and password.")
                    else:
                        try:
                            data = api_login(email, pw)
                            if data.get("needsVerification"):
                                # account exists but email unverified → route to code flow
                                st.session_state.pending_email = data.get("email", email)
                                st.session_state.dev_code = data.get("devCode", "")
                                st.info("Your email isn't verified yet — check the **Email Code** tab.")
                            else:
                                _finish_login(data)
                                st.rerun()
                        except Exception as e:
                            st.error(f"Sign-in failed: {e}")

            # ── Create Account ──
            with tab_up:
                st.markdown("##### Create your account")
                st.caption("Join ESGIntel — AI-powered ESG due diligence")
                with st.form("signup_form"):
                    n1, n2 = st.columns(2)
                    with n1:
                        first = st.text_input("First name", placeholder="Swikrit")
                    with n2:
                        last  = st.text_input("Last name", placeholder="Sharma")
                    su_email = st.text_input("Work email", placeholder="you@firm.com")
                    su_pw    = st.text_input("Password (min. 8 characters)", type="password")
                    su_pw2   = st.text_input("Confirm password", type="password")
                    su_co    = st.text_input("Company / Fund name", placeholder="Apex Capital Management")
                    su_role  = st.selectbox("Your role", ROLE_OPTIONS)
                    terms    = st.checkbox("I agree to the Terms of Service and Privacy Policy")
                    su_submit = st.form_submit_button("Create Account →", type="primary", use_container_width=True)
                if su_submit:
                    if not first or not last:
                        st.error("Enter your first and last name.")
                    elif not su_email:
                        st.error("Enter your email address.")
                    elif len(su_pw) < 8:
                        st.error("Password must be at least 8 characters.")
                    elif su_pw != su_pw2:
                        st.error("Passwords do not match.")
                    elif not terms:
                        st.error("Please accept the Terms of Service to continue.")
                    else:
                        try:
                            role = "" if su_role == "Select role…" else su_role
                            data = api_signup(first, last, su_email, su_pw, su_co, role)
                            # Backend emails a 6-digit code; if it also returns a token, we're done.
                            if data.get("token") or data.get("access_token"):
                                _finish_login(data)
                                st.rerun()
                            else:
                                st.session_state.pending_email = data.get("email", su_email)
                                st.session_state.dev_code = data.get("devCode", "")
                                st.success("Account created! Enter the 6-digit code on the **Email Code** tab.")
                        except Exception as e:
                            st.error(f"Sign-up failed: {e}")

            # ── Email Code (send + verify) ──
            with tab_code:
                st.markdown("##### Sign in with an email code")
                st.caption("We'll email you a 6-digit code — no password required.")

                code_email = st.text_input("Email address", key="code_email",
                                           value=st.session_state.pending_email,
                                           placeholder="you@firm.com")
                if st.button("Send code →", use_container_width=True, key="send_code_btn"):
                    if not code_email:
                        st.error("Enter your email address.")
                    else:
                        try:
                            data = api_send_code(code_email, purpose="login")
                            st.session_state.pending_email = code_email
                            st.session_state.dev_code = data.get("devCode", "")
                            st.success(f"Code sent to {code_email}.")
                        except Exception as e:
                            st.error(f"Could not send code: {e}")

                if st.session_state.dev_code:
                    st.info(f"🔧 Dev mode (email not configured): your code is **{st.session_state.dev_code}**")

                if st.session_state.pending_email:
                    with st.form("verify_form"):
                        code = st.text_input("Verification code", max_chars=6, placeholder="123456")
                        vsubmit = st.form_submit_button("Verify & Continue →", type="primary",
                                                        use_container_width=True)
                    if vsubmit:
                        if not code or len(code.strip()) != 6:
                            st.error("Enter the 6-digit code.")
                        else:
                            try:
                                data = api_verify_code(st.session_state.pending_email,
                                                       code.strip(), purpose="login")
                                _finish_login(data)
                                st.session_state.dev_code = ""
                                st.rerun()
                            except Exception as e:
                                st.error(f"Verification failed: {e}")

            # ── OAuth providers ──
            st.markdown("<div style='text-align:center;color:#94a3b8;font-size:12px;margin:6px 0'>or continue with</div>",
                        unsafe_allow_html=True)
            o1, o2 = st.columns(2)
            with o1:
                st.link_button("Google", f"{API_BASE}/api/auth/google/login", use_container_width=True)
            with o2:
                st.link_button("X (Twitter)", f"{API_BASE}/api/auth/x/login", use_container_width=True)

        st.markdown("""
        <div class="auth-badge-row">
          <span class="auth-badge">🔒 SOC 2 Type II</span>
          <span class="auth-badge">🇪🇺 GDPR</span>
          <span class="auth-badge">🛡 End-to-end encrypted</span>
          <span class="auth-badge">ISO 27001</span>
        </div>
        """, unsafe_allow_html=True)

        # ── Developer access (local-only convenience, clearly not a real login) ──
        with st.expander("🔧 Developer access (local only)"):
            st.caption("Skips authentication for local development. Not a real sign-in — "
                       "no user is created and protected backend calls may 403.")
            if st.button("Continue without login (dev mode)", use_container_width=True):
                st.session_state.auth_token = "dev-local-token"
                st.session_state.user = {
                    "displayName": "Local Developer", "firstName": "Local",
                    "email": "dev@localhost", "company": "ESGIntel Dev", "isAdmin": True,
                }
                st.rerun()


# ──────────────────────────────────────────────────────────────────────────────
# Topbar — persistent global shell on every authenticated page
# ──────────────────────────────────────────────────────────────────────────────
def render_topbar():
    user   = st.session_state.get("user") or {}
    result = st.session_state.get("result")
    co     = (result or {}).get("company", {})
    scores = (result or {}).get("esg_scores", {})

    name  = (user.get("displayName")
             or f"{user.get('firstName','')} {user.get('lastName','')}".strip()
             or (user.get("email", "") or "User").split("@")[0]
             or "User")
    email = user.get("email", "")
    is_admin = bool(user.get("isAdmin") or user.get("is_admin")
                    or (user.get("role", "") == "admin"))

    left, chips, conf, prof = st.columns([1.6, 3.2, 1.0, 1.3])

    with left:
        st.markdown("""
        <div class="topbar-brand">
          <div class="logo">🌱</div><div class="name">ESGIntel</div>
        </div>
        """, unsafe_allow_html=True)

    with chips:
        if result and co.get("name"):
            cov  = (result or {}).get("data_coverage", {})
            docs = cov.get("documents_analyzed", 0)
            year = cov.get("reporting_year", "—")
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
              <span class="company-chip">{co.get('name','')}</span>
              <span class="nace-chip">NACE {co.get('nace_code','—')} · {co.get('sector','—')}</span>
              <span class="topbar-meta">{docs} docs · Year {year}</span>
            </div>
            """, unsafe_allow_html=True)

    with conf:
        if result and scores.get("confidence") is not None:
            c = scores.get("confidence", 0)
            col = "#2f855a" if c >= 70 else ("#d17f1e" if c >= 40 else "#e24141")
            st.markdown(f"""
            <div class="conf-badge">
              <span class="k">Confidence</span>
              <span style="color:{col};font-weight:800">{c}%</span>
            </div>
            """, unsafe_allow_html=True)

    with prof:
        with st.popover(f"👤 {name}", use_container_width=True):
            st.markdown(f"**{name}**")
            if email:
                st.caption(email)
            st.divider()
            if st.button("👤 Account", key="tb_account", use_container_width=True):
                st.session_state.page = "account"
                st.rerun()
            if is_admin:
                if st.button("⚙️ Admin", key="tb_admin", use_container_width=True):
                    st.session_state.page = "admin"
                    st.rerun()
            st.divider()
            if st.button("↩ Log out", key="tb_logout", use_container_width=True, type="primary"):
                tok = st.session_state.get("auth_token")
                if tok and tok != "dev-local-token":
                    api_logout(tok)
                st.session_state.auth_token = None
                st.session_state.user = None
                st.session_state.pending_email = ""
                st.session_state.dev_code = ""
                st.rerun()


# ──────────────────────────────────────────────────────────────────────────────
# Sidebar navigation
# ──────────────────────────────────────────────────────────────────────────────
def render_sidebar():
    with st.sidebar:
        st.markdown("""
        <div style="display:flex;align-items:center;gap:10px;padding:6px 0 18px">
          <div style="width:32px;height:32px;background:#3154ff;border-radius:8px;display:flex;
               align-items:center;justify-content:center;font-size:16px;">🌱</div>
          <div>
            <div style="font-weight:800;font-size:16px;color:#102a43;letter-spacing:-.3px">ESGIntel</div>
            <div style="font-size:10px;font-weight:500;color:#627d98;text-transform:uppercase;letter-spacing:.5px">Due Diligence Platform</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        has_result = st.session_state.result is not None

        # New Assessment button (always visible)
        if st.button("＋  New Assessment", use_container_width=True, type="primary"):
            st.session_state.page        = "input"
            st.session_state.step        = 1
            st.session_state.analysis_id = None
            st.session_state.result      = None
            st.rerun()

        st.markdown("---")

        if has_result:
            co = st.session_state.result.get("company", {})
            scores = st.session_state.result.get("esg_scores", {})
            st.markdown(f"""
            <div class="card-box" style="margin-bottom:12px">
              <div style="font-size:11px;color:#627d98;text-transform:uppercase;font-weight:700;letter-spacing:.5px">Active Assessment</div>
              <div style="font-size:14px;font-weight:800;color:#102a43;margin-top:4px">{co.get("name","—")}</div>
              <div style="font-size:12px;color:#627d98;margin-top:2px">NACE {co.get("nace_code","—")} · {co.get("country","—")}</div>
              <div style="margin-top:10px;display:flex;gap:6px;flex-wrap:wrap">
                <span class="{score_badge_class(scores.get('overall'))}">ESG {scores.get('overall','—')}</span>
                <span class="{score_badge_class(scores.get('environmental'))}">E:{scores.get('environmental','—')}</span>
                <span class="{score_badge_class(scores.get('social'))}">S:{scores.get('social','—')}</span>
                <span class="{score_badge_class(scores.get('governance'))}">G:{scores.get('governance','—')}</span>
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("**ANALYSIS**")
            pages_main = {
                "📊 Dashboard":             "dashboard",
                "🏢 Company Profile":       "profile",
                "⚖️ Materiality Assessment": "materiality",
            }
            for label, key in pages_main.items():
                active = st.session_state.page == key
                if st.button(label, key=f"nav_{key}", use_container_width=True,
                             type="secondary" if not active else "primary"):
                    st.session_state.page = key
                    st.rerun()

            st.markdown("**ESG RISK**")
            pages_risk = {
                "🔴 ESG Risk Overview":     "esg_risk",
                "🗂️ Risk Register":         "risk_register",
                "🌿 Environmental":         "environmental",
                "👥 Social":               "social",
                "🏛️ Governance":            "governance",
            }
            for label, key in pages_risk.items():
                active = st.session_state.page == key
                if st.button(label, key=f"nav_{key}", use_container_width=True,
                             type="secondary" if not active else "primary"):
                    st.session_state.page = key
                    st.rerun()

            st.markdown("**CLIMATE & ACTIONS**")
            pages_action = {
                "🌡️ Climate Risk":          "climate",
                "💬 ESG Engagement":        "engagement",
                "📄 Report Generator":      "reports",
            }
            for label, key in pages_action.items():
                active = st.session_state.page == key
                if st.button(label, key=f"nav_{key}", use_container_width=True,
                             type="secondary" if not active else "primary"):
                    st.session_state.page = key
                    st.rerun()

        else:
            st.markdown("""
            <div style="font-size:13px;color:#627d98;padding:8px 0">
              Start a new assessment to unlock all analysis modules.
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # Past analyses
        st.markdown("**PAST ANALYSES**")
        try:
            past = api_list()
            if past:
                for item in past[:5]:
                    status_icon = "✅" if item["status"] == "complete" else ("⏳" if item["status"] == "processing" else "❌")
                    if st.button(
                        f"{status_icon} {item['company_name'][:22]}",
                        key=f"past_{item['id']}",
                        use_container_width=True
                    ):
                        if item["status"] == "complete":
                            full = api_result(item["id"])
                            st.session_state.analysis_id = item["id"]
                            st.session_state.result = full.get("result", full)
                            st.session_state.page = "dashboard"
                            st.rerun()
            else:
                st.markdown("<span style='font-size:12px;color:#627d98'>No past analyses</span>",
                            unsafe_allow_html=True)
        except Exception:
            st.markdown("<span style='font-size:12px;color:#e24141'>⚠ Backend offline</span>",
                        unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# Page: New Assessment (Input form + upload + run)
# ──────────────────────────────────────────────────────────────────────────────
def page_input():
    st.markdown('<div class="page-title">New ESG Assessment</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Complete mandatory inputs and upload supporting documents to begin.</div>', unsafe_allow_html=True)

    # Step indicator
    step = st.session_state.step
    step_labels = ["Company Inputs", "Upload Documents", "Data Validation", "Run Assessment"]
    cols = st.columns(len(step_labels))
    for i, (col, label) in enumerate(zip(cols, step_labels), 1):
        with col:
            color  = "#3154ff" if i == step else ("#2f855a" if i < step else "#cbd5e0")
            weight = "800" if i == step else "600"
            st.markdown(f"""
            <div style="text-align:center">
              <div style="width:32px;height:32px;border-radius:50%;background:{color};
                   color:white;font-size:13px;font-weight:800;line-height:32px;
                   margin:0 auto 6px">{'✓' if i < step else i}</div>
              <div style="font-size:12px;font-weight:{weight};color:{color if i <= step else '#627d98'}">{label}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Step 1: Company Inputs ──
    if step == 1:
        with st.form("step1_form"):
            c1, c2 = st.columns(2)
            with c1:
                st.markdown('#### Mandatory Inputs &nbsp; <span class="badge-red">REQUIRED</span>',
                            unsafe_allow_html=True)
                company_name = st.text_input("Company Name *",
                    value=st.session_state.company_name,
                    placeholder="e.g. Acme Manufacturing S.p.A.",
                    help="Used for external research, news monitoring, regulatory databases, ESG controversies.")

                nace_opts = NACE_OPTIONS
                nace_idx  = next((i for i, o in enumerate(nace_opts)
                                  if o.startswith(st.session_state.nace_code)), 9)
                nace_sel  = st.selectbox("NACE Sector Code *", nace_opts, index=nace_idx,
                    help="Primary driver for material ESG topics per ESRS, ISSB, SASB, GRI sector standards.")

                website_url = st.text_input("Company Website *",
                    value=st.session_state.website_url,
                    placeholder="https://company.com",
                    help="Used for public intelligence assessment when no documents are uploaded.")

            with c2:
                st.markdown("#### Company Context")
                country_idx = COUNTRIES.index(st.session_state.country) if st.session_state.country in COUNTRIES else 11
                country   = st.selectbox("Country / Region *", COUNTRIES, index=country_idx,
                    help="Primary operating jurisdiction — drives regulatory exposure (CSRD, SFDR) and country-level ESG risk.")

                emp_idx   = EMPLOYEE_RANGES.index(st.session_state.employees) if st.session_state.employees in EMPLOYEE_RANGES else 3
                employees = st.selectbox("Employee Range *", EMPLOYEE_RANGES, index=emp_idx,
                    help="Workforce size — informs social materiality, reporting thresholds, and peer benchmarking.")

                rev_idx   = REVENUE_RANGES.index(st.session_state.revenue) if st.session_state.revenue in REVENUE_RANGES else 4
                revenue   = st.selectbox("Annual Revenue (EUR) *", REVENUE_RANGES, index=rev_idx,
                    help="Annual turnover — used for intensity metrics (e.g. tCO2e/€m revenue) and size-adjusted peer comparison.")

                st.markdown("""
                <div class="card-box" style="margin-top:14px">
                  <div style="font-size:12px;font-weight:700;color:#627d98;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px">Assessment Mode</div>
                  <div style="font-size:13px;color:#334e68">
                    📄 <strong>Document mode</strong> — if you upload PDFs/reports<br>
                    🌐 <strong>Public intelligence mode</strong> — if you only provide a website URL
                  </div>
                </div>
                """, unsafe_allow_html=True)

            submitted = st.form_submit_button("Continue to Upload →", type="primary", use_container_width=True)
            if submitted:
                if not company_name:
                    st.error("Company name is required.")
                else:
                    st.session_state.company_name = company_name
                    st.session_state.nace_code    = nace_sel.split(" — ")[0]
                    st.session_state.website_url  = website_url
                    st.session_state.country      = country
                    st.session_state.employees    = employees
                    st.session_state.revenue      = revenue
                    st.session_state.step         = 2
                    st.rerun()

    # ── Step 2: Upload Documents ──
    elif step == 2:
        st.markdown(f"#### Uploading for: **{st.session_state.company_name}**")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("##### 📄 ESG Documents (Optional)")
            st.caption("PDF or TXT — policies, sustainability reports, annual reports, TCFD disclosures")
            uploaded_files = st.file_uploader(
                "Drop documents here",
                accept_multiple_files=True,
                type=["pdf", "txt"],
                key="doc_uploader",
                label_visibility="collapsed",
            )
            if uploaded_files:
                for f in uploaded_files:
                    st.markdown(f"✅ `{f.name}` ({round(f.size/1024, 1)} KB)")
            else:
                st.info("No documents uploaded — will run in Public Intelligence mode using the company website.")

        with c2:
            st.markdown("##### 📊 Quantitative ESG Data (Optional)")
            st.caption("Excel, CSV, PDF — GHG emissions, energy, water, waste, safety, diversity")
            data_files = st.file_uploader(
                "Drop data files here",
                accept_multiple_files=True,
                type=["pdf", "csv", "xlsx", "xls"],
                key="data_uploader",
                label_visibility="collapsed",
            )

            st.markdown("""
            <div class="card-box" style="margin-top:12px">
              <div style="font-size:12px;font-weight:700;color:#2563eb;margin-bottom:6px">What AI extracts:</div>
              <div style="font-size:12px;color:#334e68;line-height:1.8">
                🏭 GHG emissions (Scope 1, 2, 3)<br>
                ⚡ Energy consumption & intensity<br>
                💧 Water withdrawal & stress<br>
                🦺 LTIFR, TRIR, fatalities<br>
                👩 Board diversity & gender pay<br>
                ♻️ Waste generation & recycling<br>
                📋 Policy maturity & assurance
              </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")
        bc1, bc2, bc3 = st.columns([1, 1, 1])
        with bc1:
            if st.button("← Back", use_container_width=True):
                st.session_state.step = 1
                st.rerun()
        with bc3:
            if st.button("Continue to Validation →", type="primary", use_container_width=True):
                st.session_state["queued_files"] = uploaded_files or []
                # invalidate any prior validation result so step 3 re-runs freshly
                st.session_state.validation_rows = None
                st.session_state.validation_done = False
                st.session_state.step = 3
                st.rerun()

    # ── Step 3: Data Validation ──
    elif step == 3:
        st.markdown(f"#### Data Validation — **{st.session_state.company_name}**")
        st.caption("Review AI-parsed data before running the assessment. Confirm or correct before proceeding. "
                   "Flagged rows note where the AI made an assumption.")

        files = st.session_state.get("queued_files", [])

        # Run /api/validate once (cache the result in session_state) so it doesn't
        # re-fire on every Streamlit rerun.
        if st.session_state.validation_rows is None and not st.session_state.validation_done:
            with st.spinner("Extracting and validating data…"):
                try:
                    resp = api_validate(
                        st.session_state.company_name,
                        st.session_state.website_url,
                        files if files else None,
                    )
                    st.session_state.validation_rows = resp.get("rows", [])
                    st.session_state.validation_done = True
                except Exception as e:
                    st.session_state.validation_rows = []
                    st.session_state.validation_done = True
                    st.warning(f"Data validation could not run ({e}). "
                               "You can skip validation and proceed to the assessment.")

        rows = st.session_state.validation_rows or []
        flagged = [r for r in rows if r.get("flagged")]

        if rows:
            if flagged:
                st.markdown(f"""
                <div class="card-box" style="border-left:4px solid #d17f1e;background:#fff9f0">
                  <strong>⚠ {len(flagged)} item(s) flagged</strong> — the AI made assumptions.
                  Review before proceeding. Flags do not block scoring but may affect confidence.
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown("""
                <div class="card-box" style="border-left:4px solid #2f855a;background:#f2fbf5">
                  <strong>✓ No ambiguities flagged</strong> — parsed data looks clean.
                </div>
                """, unsafe_allow_html=True)

            df = pd.DataFrame(rows)
            show_cols = [c for c in ["metric", "value", "unit", "year", "source", "confidence"] if c in df.columns]
            disp = df.copy()
            # Prefix flagged metrics with a warning marker + show reason as its own column
            if "flagged" in disp.columns:
                disp["metric"] = disp.apply(
                    lambda r: f"⚠ {r.get('metric','')}" if r.get("flagged") else r.get("metric",""), axis=1)
                if disp["flag_reason"].astype(bool).any() if "flag_reason" in disp.columns else False:
                    show_cols = show_cols + ["flag_reason"]
            st.dataframe(disp[show_cols], use_container_width=True, hide_index=True)

            if flagged:
                st.markdown("**Flagged rows:**")
                for r in flagged:
                    st.caption(f"⚠ **{r.get('metric','')}** — {r.get('flag_reason','assumption made')}")
        else:
            st.info("📂 No parsed data available. "
                    "This is expected in Public Intelligence mode (website-only) — you can skip validation.")

        st.markdown("---")
        bc1, bc2, bc3 = st.columns([1, 1, 1.4])
        with bc1:
            if st.button("← Back", use_container_width=True):
                st.session_state.step = 2
                st.rerun()
        with bc2:
            if st.button("Skip validation", use_container_width=True,
                         help="Proceed without reviewing parsed data (useful in website-only mode)."):
                st.session_state.step = 4
                st.rerun()
        with bc3:
            if st.button("Continue to Run Assessment →", type="primary", use_container_width=True):
                st.session_state.step = 4
                st.rerun()

    # ── Step 4: Run Assessment ──
    elif step == 4:
        st.markdown(f"#### Ready to analyse: **{st.session_state.company_name}**")
        c1, c2 = st.columns([1, 1])
        with c1:
            st.markdown("""
            <div class="card-box">
              <div style="font-size:13px;font-weight:700;color:#627d98;margin-bottom:8px">ASSESSMENT INPUTS</div>
            </div>
            """, unsafe_allow_html=True)
            st.write(f"**Company:** {st.session_state.company_name}")
            st.write(f"**NACE Code:** {st.session_state.nace_code}")
            st.write(f"**Website:** {st.session_state.website_url or '—'}")
            st.write(f"**Country:** {st.session_state.country}")
            st.write(f"**Employees:** {st.session_state.employees}")
            st.write(f"**Revenue:** {st.session_state.revenue}")

            files = st.session_state.get("queued_files", [])
            if files:
                st.write(f"**Documents:** {len(files)} file(s) uploaded")
            else:
                st.info("📡 No documents — Public Intelligence mode (website scraping)")

        with c2:
            st.markdown("""
            <div class="card-box">
              <div style="font-size:13px;font-weight:700;color:#627d98;margin-bottom:8px">ANALYSIS PIPELINE</div>
              <div style="font-size:13px;color:#334e68;line-height:2">
                1. 📑 PDF discovery & report extraction<br>
                2. 🌐 Website scraping (public intelligence)<br>
                3. 🤖 AI Call 1 — Company profile & scores<br>
                4. 🤖 AI Call 2 — Deep risk, materiality & climate<br>
                5. 📊 KPI extraction (Opus structured pipeline)<br>
                6. 🏁 Results ready
              </div>
            </div>
            """, unsafe_allow_html=True)

        # ── Analysis module selection ──
        st.markdown("#### Analysis Modules")
        st.caption("Choose which modules to include in this assessment.")
        # NOTE: Only the first two toggles gate real backend behaviour today. The
        # remaining modules are UI-only stubs for now — they are shown to mirror the
        # mockup but do not yet change what /api/analyze computes.
        mc1, mc2 = st.columns(2)
        with mc1:
            st.checkbox("ESG Risk Scoring (E, S, G pillars)", value=True, key="mod_esg")           # backend-backed
            st.checkbox("Materiality Assessment", value=True, key="mod_materiality")               # backend-backed
            st.checkbox("Climate Risk & NGFS Scenarios", value=True, key="mod_climate")            # backend-backed
        with mc2:
            st.checkbox("Controversy & News Screening", value=True, key="mod_controversy")         # UI-only stub
            st.checkbox("Peer Benchmarking", value=True, key="mod_benchmark")                      # UI-only stub
            st.checkbox("SFDR / CSRD Regulatory Flags", value=False, key="mod_sfdr")               # UI-only stub

        st.markdown("---")

        if not backend_ok():
            st.error("⚠️ Backend is offline. Start the FastAPI server first: `uvicorn app:app --reload --port 8000` in the `backend/` directory.")
            return

        bc1, bc2, bc3 = st.columns([1, 2, 1])
        with bc1:
            if st.button("← Back", use_container_width=True):
                st.session_state.step = 3
                st.rerun()
        with bc2:
            run_btn = st.button("🚀 Run ESG Assessment", type="primary", use_container_width=True)

        if run_btn:
            with st.spinner("Submitting analysis…"):
                try:
                    files = st.session_state.get("queued_files", [])
                    resp  = api_submit(
                        st.session_state.company_name,
                        st.session_state.website_url,
                        st.session_state.nace_code,
                        st.session_state.country,
                        st.session_state.employees,
                        st.session_state.revenue,
                        files if files else None,
                    )
                    st.session_state.analysis_id = resp["id"]
                    st.session_state.page = "running"
                    st.rerun()
                except Exception as e:
                    st.error(f"Submission failed: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# Page: Running (progress polling)
# ──────────────────────────────────────────────────────────────────────────────
def page_running():
    aid = st.session_state.analysis_id
    st.markdown('<div class="page-title">🔄 Running ESG Assessment</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="page-sub">Company: <strong>{st.session_state.company_name}</strong> · Analysis ID: <code>{aid[:8]}…</code></div>',
                unsafe_allow_html=True)

    status_ph = st.empty()
    logs_ph   = st.empty()
    prog_ph   = st.progress(0)

    max_polls = 240  # 4 minutes
    for i in range(max_polls):
        try:
            status_data = api_status(aid)
        except Exception as e:
            status_ph.error(f"Polling error: {e}")
            time.sleep(3)
            continue

        status = status_data.get("status", "processing")
        logs   = status_data.get("logs", [])

        n_done = sum(1 for l in logs if l.get("done"))
        total  = max(len(logs), 1)
        prog_ph.progress(min(n_done / max(total, 6), 0.95))

        with logs_ph.container():
            st.markdown("**Progress log:**")
            for entry in logs[-12:]:
                icon   = "✅" if entry.get("done") else "⏳"
                elapsed = f"[{entry.get('elapsed',''):.0f}s]" if entry.get("elapsed") else ""
                cls    = "progress-done" if entry.get("done") else "progress-step"
                st.markdown(f'<div class="{cls}">{icon} {entry.get("msg","")} <span style="color:#cbd5e0;font-size:11px">{elapsed}</span></div>',
                            unsafe_allow_html=True)

        if status == "complete":
            prog_ph.progress(1.0)
            status_ph.success("✅ Assessment complete!")
            time.sleep(0.8)
            try:
                full = api_result(aid)
                st.session_state.result = full.get("result", full)
                st.session_state.page   = "dashboard"
                st.rerun()
            except Exception as e:
                st.error(f"Could not load result: {e}")
            return

        elif status == "error":
            status_ph.error(f"❌ Analysis failed: {status_data.get('error','Unknown error')}")
            if st.button("← Back to Input"):
                st.session_state.page = "input"
                st.session_state.step = 1
                st.rerun()
            return

        time.sleep(3)

    st.warning("Analysis is taking longer than expected. It may still be running — check back shortly.")


# ──────────────────────────────────────────────────────────────────────────────
# Page: Dashboard
# ──────────────────────────────────────────────────────────────────────────────
def page_dashboard():
    r      = st.session_state.result or {}
    co     = r.get("company",   {})
    scores = r.get("esg_scores",{})
    ov     = r.get("overview",  {})
    cov    = r.get("data_coverage", {})
    mode   = r.get("_mode", "document")

    # ── Header ──
    c1, c2 = st.columns([2, 1])
    with c1:
        st.markdown('<div class="page-title">📊 Dashboard</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="page-sub">{ov.get("esg_summary","")}</div>', unsafe_allow_html=True)
    with c2:
        if mode == "public_intelligence":
            st.markdown('<span class="badge-blue">🌐 Public Intelligence Mode</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="badge-green">📄 Document Mode</span>', unsafe_allow_html=True)
        docs = cov.get("documents_analyzed", 0)
        year = cov.get("reporting_year", "—")
        st.caption(f"Docs reviewed: {docs} · Reporting year: {year} · Confidence: {scores.get('confidence','—')}/100")

    # ── Analyst Snapshot banner ──
    overall = scores.get("overall", 0)
    conf    = scores.get("confidence", 0)
    tier    = "High Risk" if overall >= 70 else ("Medium Risk" if overall >= 40 else "Low Risk")
    tier_color = "#e24141" if overall >= 70 else ("#d17f1e" if overall >= 40 else "#2f855a")

    st.markdown(f"""
    <div class="card-box" style="border-left:4px solid {tier_color};display:flex;gap:24px;align-items:center;flex-wrap:wrap">
      <div>
        <div style="font-size:11px;color:#627d98;font-weight:700;text-transform:uppercase;letter-spacing:.5px">Company Quality Rating</div>
        <div style="font-size:20px;font-weight:800;color:#102a43;margin-top:2px">{co.get("name","—")}</div>
        <div style="font-size:13px;color:#627d98">{co.get("sector","—")} · {co.get("country","—")}</div>
      </div>
      <div style="margin-left:auto;text-align:center">
        <div style="font-size:11px;color:#627d98;font-weight:700;text-transform:uppercase">ESG Risk Tier</div>
        <div style="font-size:22px;font-weight:800;color:{tier_color}">{tier}</div>
      </div>
      <div style="text-align:center">
        <div style="font-size:11px;color:#627d98;font-weight:700;text-transform:uppercase">Confidence</div>
        <div style="font-size:22px;font-weight:800;color:#102a43">{conf}/100</div>
      </div>
      <div style="text-align:center">
        <div style="font-size:11px;color:#627d98;font-weight:700;text-transform:uppercase">Docs Reviewed</div>
        <div style="font-size:22px;font-weight:800;color:#102a43">{docs}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── KPI Cards ──
    st.markdown("#### ESG Risk Scores")
    kc1, kc2, kc3, kc4, kc5 = st.columns(5)
    kpi_data = [
        (kc1, "Overall ESG Score",    scores.get("overall",       0)),
        (kc2, "Environmental Risk",   scores.get("environmental", 0)),
        (kc3, "Social Risk",          scores.get("social",        0)),
        (kc4, "Governance Risk",      scores.get("governance",    0)),
        (kc5, "Data Confidence",      scores.get("confidence",    0)),
    ]
    for col, label, val in kpi_data:
        with col:
            color = score_color(val)
            st.markdown(f"""
            <div class="kpi-card">
              <div class="kpi-label">{label}</div>
              <div class="kpi-value" style="color:{color}">{val}</div>
              <div class="kpi-sub">/100</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Benchmarking vs Sector (real z-score / percentile / quartile) ──
    bench      = r.get("benchmarking", {}) or {}
    sector_ref = bench.get("sector_reference", {}) or {}
    bench_pillars = [("Overall", "overall"), ("Environmental", "environmental"),
                     ("Social", "social"), ("Governance", "governance")]
    has_bench = any(isinstance(bench.get(k), dict) and bench.get(k, {}).get("percentile") is not None
                    for _, k in bench_pillars)
    if has_bench:
        peer_label = sector_ref.get("label", "sector peers")
        n_peers    = sector_ref.get("n_peers")
        st.markdown("#### Benchmarking vs Sector")
        st.caption("Percentile is on a RISK scale — higher = riskier than peers (Q4 = worst quartile). "
                   f"Peer group: {peer_label}" + (f" · {n_peers} peers" if n_peers else ""))
        q_color = {"Q1": "#2f855a", "Q2": "#2563eb", "Q3": "#d17f1e", "Q4": "#e24141"}
        bcols = st.columns(4)
        for col, (label, key) in zip(bcols, bench_pillars):
            blk   = bench.get(key) or {}
            pct   = blk.get("percentile")
            quart = blk.get("quartile", "—")
            savg  = blk.get("sector_avg")
            col_c = q_color.get(quart, "#627d98")
            with col:
                st.markdown(f"""
                <div class="kpi-card" style="text-align:left">
                  <div class="kpi-label">{label}</div>
                  <div style="display:flex;align-items:baseline;gap:8px;margin-top:4px">
                    <div style="font-size:26px;font-weight:800;color:{col_c}">{quart}</div>
                    <div style="font-size:13px;color:#627d98;font-weight:600">{('P' + str(pct)) if pct is not None else '—'} percentile</div>
                  </div>
                  <div class="kpi-sub" style="margin-top:6px">Sector avg: {savg if savg is not None else '—'}</div>
                </div>
                """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

    # ── Climate summary + Data completeness ──
    climate  = r.get("climate", {}) or {}
    stranded = climate.get("stranded_asset_score")
    resil    = climate.get("resilience_score")
    kpis_all = r.get("kpis", []) or []
    n_kpis   = len(kpis_all)
    n_avail  = sum(1 for k in kpis_all if k.get("available", False))
    completeness = round(100 * n_avail / n_kpis) if n_kpis else None

    cl_col, dc_col = st.columns(2)
    with cl_col:
        st.markdown("#### Climate Risk Summary")
        if stranded is not None or resil is not None:
            t1, t2 = st.columns(2)
            with t1:
                sc = score_color(stranded) if stranded is not None else "#627d98"
                st.markdown(f"""
                <div class="kpi-card">
                  <div class="kpi-label">Stranded Asset Risk</div>
                  <div class="kpi-value" style="color:{sc}">{stranded if stranded is not None else '—'}</div>
                  <div class="kpi-sub">/100 · higher = worse</div>
                </div>
                """, unsafe_allow_html=True)
            with t2:
                rc = "#2f855a" if (resil or 0) >= 60 else ("#d17f1e" if (resil or 0) >= 40 else "#e24141")
                st.markdown(f"""
                <div class="kpi-card">
                  <div class="kpi-label">Climate Resilience</div>
                  <div class="kpi-value" style="color:{rc}">{resil if resil is not None else '—'}</div>
                  <div class="kpi-sub">/100 · higher = better</div>
                </div>
                """, unsafe_allow_html=True)
            if st.button("View full climate analysis →", key="dash_climate_btn"):
                st.session_state.page = "climate"
                st.rerun()
        else:
            st.info("No climate scenario data in this assessment.")

    with dc_col:
        st.markdown("#### Data Completeness")
        if completeness is not None:
            bar_color = "#2f855a" if completeness >= 66 else ("#d17f1e" if completeness >= 33 else "#e24141")
            st.markdown(f"""
            <div class="kpi-card" style="text-align:left">
              <div style="display:flex;align-items:center;gap:14px">
                <div style="font-size:34px;font-weight:800;color:{bar_color}">{completeness}%</div>
                <div style="flex:1">
                  <div style="height:8px;background:#eef3fb;border-radius:4px;overflow:hidden">
                    <div style="height:100%;width:{completeness}%;background:{bar_color};border-radius:4px"></div>
                  </div>
                  <div class="kpi-sub" style="margin-top:6px">{n_avail}/{n_kpis} KPIs evidenced · {docs} docs · Year {year}</div>
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="kpi-card" style="text-align:left">
              <div style="font-size:34px;font-weight:800;color:#627d98">{docs}</div>
              <div class="kpi-sub">documents analysed · Year {year}</div>
            </div>
            """, unsafe_allow_html=True)
        fwks_cov = cov.get("frameworks_referenced", [])
        if fwks_cov:
            st.caption("Frameworks: " + " · ".join(fwks_cov[:6]))

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Radar + Signals ──
    col_left, col_right = st.columns([1, 1])

    with col_left:
        st.markdown("#### ESG Risk Radar")
        radar_data = {
            "Environmental": scores.get("environmental", 0),
            "Social":        scores.get("social",        0),
            "Governance":    scores.get("governance",    0),
        }
        # Add climate if available
        climate_risks = r.get("climate", {})
        if climate_risks:
            phys_avg = 0
            if climate_risks.get("physical_risks"):
                phys_avg = sum(x.get("score",0) for x in climate_risks["physical_risks"]) / len(climate_risks["physical_risks"])
            if phys_avg:
                radar_data["Climate (Physical)"] = phys_avg
        st.plotly_chart(radar_chart(radar_data), use_container_width=True)

    with col_right:
        st.markdown("#### ESG Signal Feed")
        signals = r.get("signals", [])
        if signals:
            for sig in signals[:6]:
                stype = sig.get("type", "neutral")
                cls   = f"signal-{stype}"
                fw    = sig.get("framework", "")
                st.markdown(f"""
                <div class="{cls}">
                  <div class="signal-label">{sig.get("label","")}</div>
                  <div class="signal-desc">{sig.get("description","")}</div>
                  <div style="font-size:11px;color:#627d98;margin-top:3px">{fw} · Confidence: {sig.get("confidence","—")}</div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("No signals available.")

    # ── Material Topics Overview ──
    st.markdown("#### Material Topic Overview")
    topics = r.get("material_topics", [])
    if topics:
        df = pd.DataFrame(topics)
        show_cols = [c for c in ["topic", "category", "impact_score", "financial_score", "confidence", "trend"] if c in df.columns]
        st.dataframe(df[show_cols].sort_values("impact_score", ascending=False) if "impact_score" in df.columns else df[show_cols],
                     use_container_width=True, hide_index=True)
    else:
        st.info("No material topics available.")

    # ── Strengths & Gaps ──
    g1, g2 = st.columns(2)
    with g1:
        st.markdown("#### Key Strengths")
        for s in ov.get("key_strengths", []):
            st.markdown(f'✅ {s}')
    with g2:
        st.markdown("#### Key Gaps")
        for g in ov.get("key_gaps", []):
            st.markdown(f'🔴 {g}')


# ──────────────────────────────────────────────────────────────────────────────
# Page: Company Profile
# ──────────────────────────────────────────────────────────────────────────────
def page_profile():
    r   = st.session_state.result or {}
    co  = r.get("company", {})
    ov  = r.get("overview",{})
    cov = r.get("data_coverage",{})
    geo = r.get("geographic_exposure", [])
    pol = r.get("policy_maturity", [])
    disc = r.get("disclosure_gaps", [])

    st.markdown('<div class="page-title">🏢 Company Profile</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Business overview, corporate structure, and ESG maturity snapshot.</div>', unsafe_allow_html=True)

    # ── Header strip ──
    h1, h2, h3, h4 = st.columns(4)
    fields = [
        (h1, "Sector",     f"NACE {co.get('nace_code','—')} — {co.get('sector','—')}"),
        (h2, "Country",    co.get("country", "—")),
        (h3, "Employees",  co.get("employees", "—")),
        (h4, "Revenue",    co.get("revenue", "—")),
    ]
    for col, label, val in fields:
        with col:
            st.metric(label, val)

    # ── Business overview ──
    st.markdown("#### Business Overview")
    st.markdown(f"""
    <div class="card-box">
      <div style="font-size:15px;color:#334e68;line-height:1.7">{ov.get("business_summary","—")}</div>
    </div>
    """, unsafe_allow_html=True)

    # ── ESG Maturity ──
    if pol:
        st.markdown("#### ESG Maturity Snapshot")
        for p in pol:
            level = p.get("level", 0)
            label = p.get("level_label", f"Level {level}")
            area  = p.get("policy_area", "")
            color = ["#cbd5e0","#e24141","#d17f1e","#ecc94b","#68d391","#2f855a"][min(level, 5)]
            bar_w = f"{level * 20}%"
            st.markdown(f"""
            <div style="margin-bottom:12px">
              <div style="display:flex;justify-content:space-between;margin-bottom:4px">
                <span style="font-size:13px;font-weight:700;color:#102a43">{area}</span>
                <span style="font-size:12px;color:#627d98">{label}</span>
              </div>
              <div style="height:6px;background:#eef3fb;border-radius:3px;overflow:hidden">
                <div style="height:100%;width:{bar_w};background:{color};border-radius:3px;transition:width .3s"></div>
              </div>
              <div style="font-size:11.5px;color:#627d98;margin-top:3px">{p.get("observable_evidence","")}</div>
            </div>
            """, unsafe_allow_html=True)

    # ── Geographic Exposure ──
    if geo:
        st.markdown("#### Geographic Exposure")
        st.caption("Value-chain countries shaded by ESG risk level · hover for role, share and rationale.")
        geo_df = pd.DataFrame(geo)
        risk_num = {"high": 3, "medium": 2, "low": 1}
        geo_df["risk_rank"] = (geo_df.get("risk_level", pd.Series(index=geo_df.index, dtype=object))
                               .astype(str).str.lower().map(risk_num).fillna(2))

        if "iso_a3" in geo_df.columns and geo_df["iso_a3"].notna().any():
            # Choropleth — countries keyed by ISO-3166 alpha-3 code
            hover_cols = [c for c in ["role", "share", "risk_level", "climate_risk", "rationale"]
                          if c in geo_df.columns]
            fig = px.choropleth(
                geo_df,
                locations="iso_a3",
                locationmode="ISO-3",
                color="risk_rank",
                hover_name="country" if "country" in geo_df.columns else None,
                hover_data=hover_cols,
                color_continuous_scale=[[0, "#2f855a"], [0.5, "#d17f1e"], [1, "#e24141"]],
                range_color=[1, 3],
            )
            fig.update_coloraxes(showscale=False)
            fig.update_layout(
                geo=dict(showframe=False, showcoastlines=True,
                         coastlinecolor="#e2e8f0", landcolor="#f8fafc",
                         oceancolor="#e7f0ff", showocean=True, showlakes=True,
                         projection_type="natural earth"),
                height=360, margin=dict(l=0, r=0, t=0, b=0),
                paper_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(fig, use_container_width=True)
        elif "lat" in geo_df.columns and "lng" in geo_df.columns:
            # Fallback: point map when ISO-3 codes are unavailable
            geo_df["lat"] = pd.to_numeric(geo_df["lat"], errors="coerce")
            geo_df["lng"] = pd.to_numeric(geo_df["lng"], errors="coerce")
            fig = go.Figure(go.Scattergeo(
                lat=geo_df["lat"], lon=geo_df["lng"],
                text=geo_df.apply(lambda row: f"{row.get('country','?')} — {row.get('role','?')}", axis=1),
                mode="markers",
                marker=dict(size=14, color=geo_df["risk_rank"],
                            colorscale=[[0,"#2f855a"],[0.5,"#d17f1e"],[1,"#e24141"]],
                            showscale=False, line=dict(width=1, color="white")),
            ))
            fig.update_layout(
                geo=dict(showframe=False, showcoastlines=True,
                         coastlinecolor="#e2e8f0", landcolor="#f8fafc",
                         oceancolor="#e7f0ff", showocean=True, showlakes=True),
                height=340, margin=dict(l=0, r=0, t=0, b=0),
                paper_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(fig, use_container_width=True)

        # Risk legend
        st.markdown("""
        <div style="display:flex;gap:16px;font-size:12px;color:#627d98;margin:2px 0 10px">
          <span>■ <span style="color:#e24141">High risk</span></span>
          <span>■ <span style="color:#d17f1e">Medium risk</span></span>
          <span>■ <span style="color:#2f855a">Low risk</span></span>
        </div>
        """, unsafe_allow_html=True)

        show_cols = [c for c in ["country","role","share","risk_level","climate_risk","rationale"]
                     if c in geo_df.columns]
        st.dataframe(geo_df[show_cols], use_container_width=True, hide_index=True)

        # ── Applicable regulations by country ──
        countries_with_regs = [g for g in geo if g.get("regulations")]
        if countries_with_regs:
            st.markdown("#### Applicable Regulations by Country")
            st.caption("Regulatory exposure surfaced per value-chain country.")
            for g in countries_with_regs:
                regs  = g.get("regulations", []) or []
                cname = g.get("country", "—")
                with st.expander(f"{cname}  ·  {g.get('role','—')}  ·  {len(regs)} regulation(s)"):
                    for reg in regs:
                        if isinstance(reg, dict):
                            rn = (reg.get("law") or reg.get("name") or reg.get("regulation")
                                  or reg.get("title") or "—")
                            rd = (reg.get("scope") or reg.get("description")
                                  or reg.get("detail") or reg.get("relevance") or "")
                            st.markdown(f"- **{rn}**" + (f" — {rd}" if rd else ""))
                        else:
                            st.markdown(f"- {reg}")
                    if g.get("rationale"):
                        st.caption(g.get("rationale"))

    # ── Disclosure Gaps ──
    if disc:
        st.markdown("#### Disclosure Gaps")
        for gap in disc:
            sev = gap.get("severity", "Medium")
            st.markdown(f"""
            <div class="card-box" style="border-left:3px solid {'#e24141' if sev=='Critical' else '#d17f1e' if sev in ('High','Medium') else '#2f855a'}">
              <div style="display:flex;justify-content:space-between">
                <div style="font-size:13px;font-weight:700">{gap.get("area","")}</div>
                {severity_badge(sev)}
              </div>
              <div style="font-size:12px;color:#627d98;margin-top:4px">{gap.get("type","")}</div>
              <div style="font-size:13px;color:#334e68;margin-top:6px">{gap.get("explanation","")}</div>
              <div style="font-size:12px;color:#2563eb;margin-top:4px">💡 {gap.get("recommendation","")}</div>
            </div>
            """, unsafe_allow_html=True)

    # ── Data Coverage ──
    st.markdown("#### Data Coverage & Completeness")
    d1, d2, d3, d4 = st.columns(4)
    with d1:
        st.metric("Documents Analyzed", cov.get("documents_analyzed", 0))
    with d2:
        st.metric("Reporting Year", cov.get("reporting_year", "—"))
    with d3:
        st.metric("Assurance Level", cov.get("assurance", "Unknown"))
    with d4:
        st.metric("Scope 3 Disclosed", "Yes" if cov.get("scope3_disclosed") else "No")

    fwks = cov.get("frameworks_referenced", [])
    if fwks:
        st.markdown("**Frameworks referenced:** " + " · ".join(f"`{f}`" for f in fwks))


# ──────────────────────────────────────────────────────────────────────────────
# Page: Materiality Assessment
# ──────────────────────────────────────────────────────────────────────────────
def page_materiality():
    r      = st.session_state.result or {}
    topics = r.get("material_topics", [])

    st.markdown('<div class="page-title">⚖️ Materiality Assessment</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Dual materiality (financial + impact). Sources: SASB sector standards, ESRS, GRI 3.</div>', unsafe_allow_html=True)

    if not topics:
        st.info("No material topics in this assessment.")
        return

    overrides = st.session_state.materiality_overrides

    def _eff(t):
        """Shallow copy of topic t with any session-only score override applied."""
        ov = overrides.get(t.get("topic", ""), {})
        et = dict(t)
        if "impact_score" in ov:
            et["impact_score"] = ov["impact_score"]
        if "financial_score" in ov:
            et["financial_score"] = ov["financial_score"]
        et["_overridden"] = bool(ov)
        return et

    effective = [_eff(t) for t in topics]

    if overrides:
        st.caption(f"✏️ {len(overrides)} topic(s) have manual score overrides — "
                   "overrides apply to this session only and are not saved to the backend.")

    # Filter controls
    cats = ["All"] + sorted({t.get("category","") for t in topics if t.get("category")})
    cat_filter = st.selectbox("Filter by category", cats, key="mat_cat_filter")
    filtered = effective if cat_filter == "All" else [t for t in effective if t.get("category") == cat_filter]

    # ── Dual materiality matrix ──
    st.markdown("#### Dual Materiality Matrix")
    fig = materiality_matrix(filtered)
    if fig:
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("""
    <div style="display:flex;gap:16px;font-size:12px;color:#627d98;margin-bottom:8px">
      <span>● <span style="color:#2f855a">Environmental</span></span>
      <span>● <span style="color:#2563eb">Social</span></span>
      <span>● <span style="color:#d17f1e">Governance</span></span>
      <span style="margin-left:auto">Confidence: 🟢 High · 🟡 Medium · 🔴 Low</span>
    </div>
    """, unsafe_allow_html=True)

    # ── Topic table ──
    conf_dot = {"high": "🟢", "medium": "🟡", "low": "🔴"}
    st.markdown("#### Material Topic Detail")
    st.caption("Expand a topic to review evidence or override its impact / financial materiality scores. "
               "Overrides apply to this session only.")
    for t in sorted(filtered, key=lambda x: -(x.get("impact_score",0) + x.get("financial_score",0))):
        name = t.get("topic", "—")
        imp = t.get("impact_score",   0)
        fin = t.get("financial_score",0)
        cat = t.get("category", "")
        conf = (t.get("confidence") or "").lower()
        dot = conf_dot.get(conf, "⚪")
        flag = " ✏️" if t.get("_overridden") else ""
        cat_color = {"environmental":"#2f855a","social":"#2563eb","governance":"#d17f1e"}.get(cat,"#627d98")

        with st.expander(f"{dot} {name}{flag}  —  Impact: {imp}/5  ·  Financial: {fin}/5"):
            cc1, cc2, cc3 = st.columns(3)
            with cc1:
                st.markdown(f"**Category:** <span style='color:{cat_color}'>{cat.title()}</span>", unsafe_allow_html=True)
                st.markdown(f"**Trend:** {(t.get('trend') or '—').title()}")
                st.markdown(f"**Confidence:** {confidence_badge(t.get('confidence'))}", unsafe_allow_html=True)
            with cc2:
                st.markdown("**Rationale:**")
                st.markdown(t.get("rationale", "—"))
                if t.get("financial_rationale"):
                    st.markdown("**Financial rationale:**")
                    st.markdown(t.get("financial_rationale"))
            with cc3:
                if t.get("evidence"):
                    st.markdown(f"""
                    <div class="evidence-box">
                      <div class="evidence-source">Evidence</div>
                      <div class="evidence-text">{t.get('evidence','')}</div>
                    </div>
                    """, unsafe_allow_html=True)
                fwks = t.get("frameworks", [])
                if fwks:
                    st.markdown("**Frameworks:** " + " · ".join(f"`{f}`" for f in fwks))

            # ── Override scores (session-only) ──
            st.markdown("---")
            st.markdown("**Override scores** &nbsp;<span class='badge-gray'>session only</span>",
                        unsafe_allow_html=True)
            orig = next((o for o in topics if o.get("topic") == name), {})
            orig_imp = float(orig.get("impact_score", 0) or 0)
            orig_fin = float(orig.get("financial_score", 0) or 0)
            oc1, oc2, oc3 = st.columns([1, 1, 1])
            with oc1:
                new_imp = st.number_input("Impact materiality", min_value=0.0, max_value=5.0, step=0.5,
                                          value=orig_imp, key=f"ov_imp_{name}")
            with oc2:
                new_fin = st.number_input("Financial materiality", min_value=0.0, max_value=5.0, step=0.5,
                                          value=orig_fin, key=f"ov_fin_{name}")
            with oc3:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("↺ Reset", key=f"ov_reset_{name}", use_container_width=True):
                    overrides.pop(name, None)
                    st.session_state.pop(f"ov_imp_{name}", None)
                    st.session_state.pop(f"ov_fin_{name}", None)
                    st.rerun()
            # Persist an override only when the values actually diverge from the original
            if new_imp != orig_imp or new_fin != orig_fin:
                if overrides.get(name) != {"impact_score": new_imp, "financial_score": new_fin}:
                    overrides[name] = {"impact_score": new_imp, "financial_score": new_fin}
                    st.rerun()
            elif name in overrides:
                overrides.pop(name, None)
                st.rerun()


# ──────────────────────────────────────────────────────────────────────────────
# Shared: Risk list renderer (used by E/S/G/overview pages)
# ──────────────────────────────────────────────────────────────────────────────
def render_risk_list(risks, category_filter=None):
    filtered = risks
    if category_filter:
        filtered = [r for r in risks if r.get("category") == category_filter]

    if not filtered:
        st.info("No risks found in this category.")
        return

    # Sort by score desc
    filtered = sorted(filtered, key=lambda x: -(x.get("score") or 0))

    for risk in filtered:
        score = risk.get("score", 0)
        sev   = risk.get("severity", "Medium")
        cat   = risk.get("category", "")
        color = score_color(score)

        with st.expander(f"{'🔴' if score >= 70 else '🟡' if score >= 40 else '🟢'} {risk.get('name','—')}  ·  Score: {score}  ·  {sev}"):
            r1, r2, r3 = st.columns([1, 2, 1])
            with r1:
                st.markdown(f"**Score:** <span style='color:{color};font-size:22px;font-weight:800'>{score}</span>/100", unsafe_allow_html=True)
                st.markdown(f"**Severity:** {severity_badge(sev)}", unsafe_allow_html=True)
                st.markdown(f"**Category:** {cat.title()}")
                st.markdown(f"**Framework:** `{risk.get('framework','—')}`")
                if risk.get("confidence"):
                    st.markdown(f"**AI Confidence:** {confidence_badge(risk['confidence'])}", unsafe_allow_html=True)

            with r2:
                st.markdown("**Detail:**")
                st.markdown(risk.get("detail", "—"))
                if risk.get("recommendation"):
                    st.markdown("**Recommendation:**")
                    st.markdown(f"> 💡 {risk.get('recommendation')}")

            with r3:
                evidence = risk.get("evidence", [])
                if isinstance(evidence, list) and evidence:
                    st.markdown("**Evidence:**")
                    for ev in evidence[:3]:
                        src  = ev.get("source","") if isinstance(ev, dict) else str(ev)
                        text = ev.get("text","")   if isinstance(ev, dict) else ""
                        conf = ev.get("confidence","") if isinstance(ev, dict) else ""
                        st.markdown(f"""
                        <div class="evidence-box">
                          <div class="evidence-source">{src}</div>
                          <div class="evidence-text">{text}</div>
                          {'<div class="evidence-conf">Confidence: ' + conf + '</div>' if conf else ''}
                        </div>
                        """, unsafe_allow_html=True)

                kpis = risk.get("kpis", [])
                if kpis:
                    st.markdown("**KPIs:**")
                    for kpi in kpis[:3]:
                        val  = kpi.get("value","—")
                        unit = kpi.get("unit","")
                        st.markdown(f"- **{kpi.get('metric','')}:** {val} {unit}")


# ──────────────────────────────────────────────────────────────────────────────
# Page: ESG Risk Overview
# ──────────────────────────────────────────────────────────────────────────────
def page_esg_risk():
    r      = st.session_state.result or {}
    risks  = r.get("risks", [])
    scores = r.get("esg_scores", {})

    st.markdown('<div class="page-title">🔴 ESG Risk Overview</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">All identified ESG risks scored, categorised, and evidence-backed. '
                'Click a pillar below to drill into its detailed analysis.</div>', unsafe_allow_html=True)

    # Score summary — each pillar cell drills into its dedicated page
    bench = r.get("benchmarking", {}) or {}
    s1, s2, s3 = st.columns(3)
    for col, label, key, page_key in [(s1,"Environmental", "environmental", "environmental"),
                                       (s2,"Social",        "social",        "social"),
                                       (s3,"Governance",    "governance",    "governance")]:
        val = scores.get(key, 0)
        blk = bench.get(key) or {}
        pct, quart = blk.get("percentile"), blk.get("quartile")
        with col:
            st.plotly_chart(gauge_chart(val, label), use_container_width=True)
            if pct is not None:
                st.caption(f"{quart or '—'} · P{pct} vs sector")
            if st.button(f"View {label} →", key=f"drill_{page_key}", use_container_width=True):
                st.session_state.page = page_key
                st.rerun()

    # Risk bar chart
    st.markdown("#### Risk Scores Overview")
    fig = risk_bar_chart(risks)
    if fig:
        st.plotly_chart(fig, use_container_width=True)

    # Full risk list
    hd1, hd2 = st.columns([3, 1])
    with hd1:
        st.markdown("#### All ESG Risks")
    with hd2:
        if st.button("🗂️ View Full Risk Register →", use_container_width=True, key="goto_rr"):
            st.session_state.page = "risk_register"
            st.rerun()
    render_risk_list(risks)


# ──────────────────────────────────────────────────────────────────────────────
# Shared helpers for the pillar subtabs + dedicated Risk Register page
# ──────────────────────────────────────────────────────────────────────────────
def _fmt(val, fallback="—"):
    """Return val unless it's empty/None, else the fallback string."""
    return val if val not in (None, "", []) else fallback

# Keyword → pillar map, used to classify un-tagged policy_maturity entries.
_PILLAR_KW = {
    "environmental": ["environ", "climate", "emission", "energy", "water", "waste",
                      "biodivers", "pollution", "ghg", "carbon", "circular", "nature"],
    "social":        ["social", "health", "safety", "labour", "labor", "human right",
                      "divers", "employee", "community", "workforce", "training",
                      "supply chain", "supplier", "product safety"],
    "governance":    ["govern", "ethic", "board", "complian", "corrupt", "brib",
                      "risk manage", "data", "cyber", "tax", "remunerat", "whistle",
                      "transparen", "audit", "conduct"],
}

def _classify_pillar(text):
    t = (text or "").lower()
    for pillar, kws in _PILLAR_KW.items():
        if any(k in t for k in kws):
            return pillar
    return None

def _collect_evidence(items):
    """Flatten the evidence[] arrays from a list of risks/topics into table rows."""
    out = []
    for it in items:
        origin = it.get("name") or it.get("topic", "") or ""
        for e in (it.get("evidence") or []):
            if isinstance(e, dict):
                out.append({"Source": e.get("source", ""), "Snippet": e.get("text", ""),
                            "Confidence": e.get("confidence", ""), "From": origin})
            elif e:
                out.append({"Source": str(e), "Snippet": "", "Confidence": "", "From": origin})
    return out

def resilience_gauge(value, title="Climate Resilience"):
    """Gauge where HIGHER = better (resilience is inverted vs the risk gauges)."""
    col = "#2f855a" if value >= 60 else ("#d17f1e" if value >= 35 else "#e24141")
    fig = go.Figure(go.Indicator(
        mode="gauge+number", value=value,
        title={"text": title, "font": {"size": 13, "color": "#627d98"}},
        number={"font": {"size": 28, "color": col, "family": "Inter"}},
        gauge={"axis": {"range": [0, 100], "tickwidth": 1}, "bar": {"color": col},
               "bgcolor": "white", "borderwidth": 0,
               "steps": [{"range": [0, 35], "color": "#fff2f2"},
                         {"range": [35, 60], "color": "#fff4de"},
                         {"range": [60, 100], "color": "#e9f7ee"}]},
    ))
    fig.update_layout(height=180, margin=dict(l=20, r=20, t=30, b=0),
                      paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                      font=dict(family="Inter, sans-serif"))
    return fig

def render_risk_register(risks, key_prefix="rr", show_pillar=True):
    """Summary table + expandable rows for a set of risks. The key_driver / horizon /
    financial_impact columns are hidden entirely when NO risk in the set carries them
    (older cached analyses ran before that schema change)."""
    if not risks:
        st.info("No risks in this selection.")
        return

    risks = sorted(risks, key=lambda x: -(x.get("score") or 0))
    has_driver  = any(rk.get("key_driver")      for rk in risks)
    has_horizon = any(rk.get("horizon")         for rk in risks)
    has_fin     = any(rk.get("financial_impact") for rk in risks)

    rows = []
    for rk in risks:
        row = {"Risk": rk.get("name", "—")}
        if show_pillar:
            row["Pillar"] = (rk.get("category", "") or "").title()
        row["Score"]    = rk.get("score", 0)
        row["Severity"] = rk.get("severity", "—")
        if has_horizon:
            row["Horizon"] = rk.get("horizon") or "—"
        if has_driver:
            row["Key Driver"] = rk.get("key_driver") or "—"
        if has_fin:
            row["Financial Impact"] = rk.get("financial_impact") or "—"
        rows.append(row)
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    if not (has_driver and has_horizon and has_fin):
        st.caption("Root-cause, horizon and financial-exposure fields populate only for analyses run "
                   "under the current schema — re-run the assessment to see them for older results.")

    for i, rk in enumerate(risks):
        score = rk.get("score", 0)
        icon  = "🔴" if score >= 70 else "🟡" if score >= 40 else "🟢"
        with st.expander(f"{icon} {rk.get('name','—')}  ·  Score {score}  ·  {rk.get('severity','—')}"):
            m1, m2 = st.columns(2)
            with m1:
                st.markdown(f"**Root cause (key driver):** {_fmt(rk.get('key_driver'), 'Not available for this analysis')}")
                st.markdown(f"**Time horizon:** {_fmt(rk.get('horizon'), 'Not available for this analysis')}")
            with m2:
                st.markdown(f"**Financial exposure:** {_fmt(rk.get('financial_impact'), 'Not available for this analysis')}")
                st.markdown(f"**Framework:** `{rk.get('framework','—')}`")
            st.markdown("**Detail:**")
            st.markdown(rk.get("detail", "—"))
            if rk.get("recommendation"):
                st.markdown(f"**Mitigation:** 💡 {rk.get('recommendation')}")
            ev = rk.get("evidence", [])
            if isinstance(ev, list) and ev:
                st.markdown("**Source / evidence:**")
                for e in ev[:3]:
                    if isinstance(e, dict):
                        st.markdown(f"""
                        <div class="evidence-box">
                          <div class="evidence-source">{e.get('source','')}</div>
                          <div class="evidence-text">{e.get('text','')}</div>
                          {'<div class="evidence-conf">Confidence: ' + str(e.get('confidence','')) + '</div>' if e.get('confidence') else ''}
                        </div>""", unsafe_allow_html=True)
                    else:
                        st.caption(str(e))


# ──────────────────────────────────────────────────────────────────────────────
# Page: Environmental / Social / Governance (shared template — 6 subtabs)
# ──────────────────────────────────────────────────────────────────────────────
def page_pillar(category: str):
    r      = st.session_state.result or {}
    risks  = r.get("risks", []) or []
    scores = r.get("esg_scores", {}) or {}

    icons  = {"environmental": "🌿", "social": "👥", "governance": "🏛️"}
    titles = {"environmental": "Environmental Risk", "social": "Social Risk", "governance": "Governance Risk"}
    icon   = icons.get(category, "")
    title  = titles.get(category, category.title())

    st.markdown(f'<div class="page-title">{icon} {title}</div>', unsafe_allow_html=True)

    score = scores.get(category, 0)
    bench = (r.get("benchmarking", {}) or {}).get(category, {}) or {}

    # ── Score header (kept above the subtabs) ──
    hc1, hc2, hc3, hc4, hc5 = st.columns([1.6, 1, 1, 1, 1])
    with hc1:
        st.plotly_chart(gauge_chart(score, title), use_container_width=True)
    with hc2:
        st.metric("Score", f"{score}/100")
    with hc3:
        st.metric("Sector Avg", bench.get("sector_avg", "—") if bench.get("sector_avg") is not None else "—")
    with hc4:
        pct = bench.get("percentile")
        st.metric("Percentile", f"P{pct}" if pct is not None else "—")
    with hc5:
        st.metric("Quartile", bench.get("quartile", "—") or "—")

    cat_risks   = [rk for rk in risks if rk.get("category") == category]
    topics_here = [t for t in (r.get("material_topics") or []) if t.get("category") == category]

    tabs = st.tabs(["Risk Register", "Quantitative Metrics", "Qualitative Findings",
                    "Benchmarking", "Evidence & Sources", "Recommended Actions"])

    # ── 1. Risk Register (pillar-filtered) ──
    with tabs[0]:
        render_risk_register(cat_risks, key_prefix=f"{category}_rr", show_pillar=False)

    # ── 2. Quantitative Metrics ──
    with tabs[1]:
        cat_kpis = [k for k in (r.get("kpis", []) or [])
                    if k.get("category") == category and k.get("available", True)]
        if cat_kpis:
            kdf = pd.DataFrame(cat_kpis)
            show_yoy = False
            if "year" in kdf.columns and "metric" in kdf.columns:
                yrs = kdf.dropna(subset=["year"]).groupby("metric")["year"].nunique()
                show_yoy = bool((yrs > 1).any())
            cols = [c for c in ["metric", "value", "unit", "year", "source", "benchmark", "percentile"]
                    if c in kdf.columns]
            st.dataframe(kdf[cols], use_container_width=True, hide_index=True)
            if not show_yoy:
                st.caption("Year-over-year deltas are shown only when a metric has more than one reporting "
                           "year — this analysis carries a single year of data (YoY: N/A).")
        else:
            st.info("No quantitative metrics were evidenced for this pillar.")

    # ── 3. Qualitative Findings ──
    with tabs[2]:
        pol      = r.get("policy_maturity", []) or []
        pol_here = [p for p in pol
                    if _classify_pillar(f"{p.get('policy_area','')} {p.get('observable_evidence','')}") == category]
        shown_any = False
        if bench.get("narrative"):
            st.markdown("##### Analyst Narrative")
            st.markdown(bench["narrative"])
            shown_any = True
        if pol_here:
            shown_any = True
            st.markdown("##### Policy & Management Maturity")
            for p in pol_here:
                level = int(p.get("level", 0) or 0)
                label = p.get("level_label", f"Level {level}")
                color = ["#cbd5e0", "#e24141", "#d17f1e", "#ecc94b", "#68d391", "#2f855a"][min(level, 5)]
                st.markdown(f"""
                <div style="margin-bottom:12px">
                  <div style="display:flex;justify-content:space-between;margin-bottom:4px">
                    <span style="font-size:13px;font-weight:700;color:#102a43">{p.get('policy_area','')}</span>
                    <span style="font-size:12px;color:#627d98">{label}</span>
                  </div>
                  <div style="height:6px;background:#eef3fb;border-radius:3px;overflow:hidden">
                    <div style="height:100%;width:{level*20}%;background:{color};border-radius:3px"></div>
                  </div>
                  <div style="font-size:11.5px;color:#627d98;margin-top:3px">{p.get('observable_evidence','')}</div>
                </div>""", unsafe_allow_html=True)
        if topics_here:
            shown_any = True
            st.markdown("##### Material Topics")
            for t in topics_here:
                st.markdown(
                    f"- **{t.get('topic','')}** — {t.get('rationale','')} "
                    f"<span class='badge-gray'>Impact {t.get('impact_score','—')} · "
                    f"Financial {t.get('financial_score','—')}</span>",
                    unsafe_allow_html=True)
        if not shown_any:
            st.info("No qualitative findings available for this pillar.")

    # ── 4. Benchmarking ──
    with tabs[3]:
        if bench:
            b1, b2, b3, b4 = st.columns(4)
            b1.metric("Score", f"{score}/100")
            b2.metric("Sector Avg", bench.get("sector_avg", "—") if bench.get("sector_avg") is not None else "—")
            b3.metric("Percentile", f"P{bench.get('percentile')}" if bench.get("percentile") is not None else "—")
            b4.metric("Quartile", bench.get("quartile", "—") or "—")
            if bench.get("z_score") is not None:
                st.caption(f"Z-score vs sector: {bench.get('z_score')} · Percentile is on a RISK scale "
                           "(higher percentile / Q4 = riskier than peers).")
            if bench.get("narrative"):
                st.markdown(bench["narrative"])
            if bench.get("peer_group"):
                st.caption(f"Peer group: {bench['peer_group']}")
        else:
            st.info("No benchmarking data available for this pillar.")

    # ── 5. Evidence & Sources ──
    with tabs[4]:
        ev_rows = _collect_evidence(cat_risks) + _collect_evidence(topics_here)
        if ev_rows:
            st.caption("Aggregated source citations behind this pillar's risks and material topics.")
            st.dataframe(pd.DataFrame(ev_rows), use_container_width=True, hide_index=True)
        else:
            st.info("No evidence citations were captured for this pillar.")

    # ── 6. Recommended Actions ──
    with tabs[5]:
        actions, seen = [], set()
        for rk in sorted(cat_risks, key=lambda x: -(x.get("score") or 0)):
            rec = (rk.get("recommendation") or "").strip()
            if not rec:
                continue
            dedupe_key = rec.lower()[:80]
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            actions.append({"Priority": rk.get("severity", "Medium"), "Risk": rk.get("name", ""),
                            "Recommended Action": rec, "Framework": rk.get("framework", "")})
        if actions:
            st.dataframe(pd.DataFrame(actions), use_container_width=True, hide_index=True)
        else:
            st.info("No standalone recommendation fields were parsed for this pillar. "
                    "Mitigation guidance is included within each risk on the Risk Register tab.")


# ──────────────────────────────────────────────────────────────────────────────
# Page: ESG Risk Register (full cross-pillar register + Excel export)
# ──────────────────────────────────────────────────────────────────────────────
def _risk_register_xlsx(risks, company_name):
    """Local openpyxl export of the LIVE risks[]. There is no backend endpoint that
    emits an xlsx from the current analysis result (/api/documents/engagement serves
    canned demo data only), so we build the workbook here from the real risks."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"
    headers = ["Risk", "Pillar", "Score", "Severity", "Horizon", "Key Driver",
               "Financial Impact", "Framework", "Detail", "Recommendation"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="3154FF")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
        c.alignment = Alignment(vertical="center")
    for rk in sorted(risks, key=lambda x: -(x.get("score") or 0)):
        ws.append([
            rk.get("name", ""), (rk.get("category", "") or "").title(), rk.get("score", ""),
            rk.get("severity", ""), rk.get("horizon", ""), rk.get("key_driver", ""),
            rk.get("financial_impact", ""), rk.get("framework", ""),
            rk.get("detail", ""), rk.get("recommendation", ""),
        ])
    for i, w in enumerate([34, 14, 8, 10, 14, 30, 18, 18, 60, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def page_risk_register():
    r     = st.session_state.result or {}
    risks = r.get("risks", []) or []
    co    = r.get("company", {}) or {}

    st.markdown('<div class="page-title">🗂️ ESG Risk Register</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Complete register of identified ESG risks across all pillars — filter, '
                'sort, drill into root cause and financial exposure, and export to Excel.</div>',
                unsafe_allow_html=True)

    if not risks:
        st.info("No risks available. Run an assessment first.")
        return

    def _n(cat):
        return sum(1 for rk in risks if rk.get("category") == cat)
    counts = {"All": len(risks), "Environmental": _n("environmental"), "Social": _n("social"),
              "Governance": _n("governance"), "Climate": _n("climate")}
    filter_opts = [f"{k} ({v})" for k, v in counts.items() if k == "All" or v > 0]

    fc1, fc2, fc3 = st.columns([2.4, 1.3, 1.1])
    with fc1:
        pick = st.radio("Pillar", filter_opts, horizontal=True, key="rr_pillar")
    with fc2:
        sort_field = st.selectbox("Sort by", ["Score", "Pillar", "Financial impact", "Horizon", "Severity"],
                                  key="rr_sortf")
    with fc3:
        sort_dir = st.selectbox("Order", ["Descending", "Ascending"], key="rr_sortd")

    chosen = pick.split(" (")[0]
    view = risks if chosen == "All" else [rk for rk in risks if rk.get("category") == chosen.lower()]

    horizon_rank = {"short-term": 1, "medium-term": 2, "long-term": 3}
    sev_rank     = {"low": 1, "medium": 2, "high": 3, "critical": 4}
    def sort_key(rk):
        if sort_field == "Score":            return rk.get("score") or 0
        if sort_field == "Pillar":           return rk.get("category") or ""
        if sort_field == "Financial impact": return (rk.get("financial_impact") or "")
        if sort_field == "Horizon":          return horizon_rank.get((rk.get("horizon") or "").lower(), 0)
        if sort_field == "Severity":         return sev_rank.get((rk.get("severity") or "").lower(), 0)
        return 0
    view = sorted(view, key=sort_key, reverse=(sort_dir == "Descending"))

    ec1, _ = st.columns([1, 3])
    with ec1:
        try:
            xls = _risk_register_xlsx(view, co.get("name", "company"))
            cname = (co.get("name", "company") or "company").replace(" ", "_").replace("/", "_")
            st.download_button("⬇ Export to Excel", data=xls,
                               file_name=f"ESGIntel_{cname}_Risk_Register.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True, key="rr_export")
        except Exception as e:
            st.caption(f"Excel export unavailable: {e}")

    st.markdown(f"#### {len(view)} risk(s) in view")
    render_risk_register(view, key_prefix="page_rr", show_pillar=True)


# ──────────────────────────────────────────────────────────────────────────────
# Page: Climate Risk
# ──────────────────────────────────────────────────────────────────────────────
def page_climate():
    r       = st.session_state.result or {}
    climate = r.get("climate", {})
    phys    = climate.get("physical_risks",   [])
    trans   = climate.get("transition_risks", [])

    st.markdown('<div class="page-title">🌡️ Climate Risk</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Physical and transition risk assessment. Frameworks: IFRS S2, TCFD, NGFS scenarios.</div>', unsafe_allow_html=True)

    # ── Headline resilience/stranded strip (real Python-computed scores) ──
    _stranded = climate.get("stranded_asset_score")
    _resil    = climate.get("resilience_score")
    if _stranded is not None or _resil is not None:
        hs1, hs2 = st.columns(2)
        with hs1:
            if _stranded is not None:
                st.metric("Stranded Asset Risk", f"{_stranded}/100", help="Higher = greater stranded-asset exposure.")
        with hs2:
            if _resil is not None:
                st.metric("Climate Resilience", f"{_resil}/100", help="Higher = more resilient.")

    tab_phys, tab_trans, tab_scen, tab_resil = st.tabs(
        ["🌊 Physical Risks", "⚡ Transition Risks", "📉 Scenario Analysis", "🛡️ Resilience Assessment"])

    def risk_card(risk, kind="physical"):
        score   = risk.get("score", 0)
        color   = score_color(score)
        meta    = risk.get("scenario") if kind == "physical" else risk.get("horizon","")
        conf    = risk.get("confidence","")
        return f"""
        <div class="card-box" style="border-left:4px solid {color}">
          <div style="display:flex;justify-content:space-between;align-items:flex-start">
            <div style="font-size:15px;font-weight:800;color:#102a43">{risk.get("name","—")}</div>
            <div style="font-size:24px;font-weight:800;color:{color}">{score}</div>
          </div>
          <div style="font-size:12px;color:#627d98;margin-top:2px">{meta} {'· Confidence: ' + conf.title() if conf else ''}</div>
          <div style="font-size:13px;color:#334e68;margin-top:8px;line-height:1.6">{risk.get("detail","—")}</div>
        </div>
        """

    with tab_phys:
        if phys:
            c1, c2 = st.columns(2)
            cols = [c1, c2]
            for i, risk in enumerate(phys):
                with cols[i % 2]:
                    st.markdown(risk_card(risk, "physical"), unsafe_allow_html=True)
        else:
            st.info("No physical risks identified.")

        # Physical risk bar
        if phys:
            df_p = pd.DataFrame(phys)[["name","score"]].copy()
            df_p["score"] = pd.to_numeric(df_p["score"], errors="coerce").fillna(0)
            df_p = df_p.sort_values("score")
            fig = go.Figure(go.Bar(
                x=df_p["score"], y=df_p["name"], orientation="h",
                marker_color=[score_color(s) for s in df_p["score"]],
            ))
            fig.update_layout(height=max(200, len(df_p)*40),
                              margin=dict(l=0,r=20,t=0,b=0),
                              paper_bgcolor="rgba(0,0,0,0)",
                              plot_bgcolor="rgba(0,0,0,0)",
                              xaxis_title="Risk Score")
            st.plotly_chart(fig, use_container_width=True)

    with tab_trans:
        if trans:
            c1, c2 = st.columns(2)
            cols = [c1, c2]
            for i, risk in enumerate(trans):
                with cols[i % 2]:
                    st.markdown(risk_card(risk, "transition"), unsafe_allow_html=True)
        else:
            st.info("No transition risks identified.")

        if trans:
            df_t = pd.DataFrame(trans)[["name","score"]].copy()
            df_t["score"] = pd.to_numeric(df_t["score"], errors="coerce").fillna(0)
            df_t = df_t.sort_values("score")
            fig2 = go.Figure(go.Bar(
                x=df_t["score"], y=df_t["name"], orientation="h",
                marker_color=[score_color(s) for s in df_t["score"]],
            ))
            fig2.update_layout(height=max(200, len(df_t)*40),
                               margin=dict(l=0,r=20,t=0,b=0),
                               paper_bgcolor="rgba(0,0,0,0)",
                               plot_bgcolor="rgba(0,0,0,0)",
                               xaxis_title="Risk Score")
            st.plotly_chart(fig2, use_container_width=True)

    # ── Scenario Analysis (NGFS-aligned, real modeled financials) ──
    with tab_scen:
        scenarios = climate.get("scenarios") or []
        if not scenarios:
            st.info("🔄 Scenario modeling isn't available for this analysis. "
                    "Re-run the assessment to generate NGFS-aligned scenario financials.")
        else:
            st.markdown("##### NGFS Scenario Analysis — Financial Impact")
            st.caption("⚠️ Modeled estimate, not audited financial guidance. Figures are Python-computed from "
                       "emissions, carbon-price trajectories and revenue exposure using NGFS scenario parameters.")
            colors = ["#2f855a", "#d17f1e", "#e24141"]
            cols = st.columns(len(scenarios))
            for i, (col, sc) in enumerate(zip(cols, scenarios)):
                c = colors[min(i, len(colors) - 1)]
                imp = sc.get("financial_impact_eur_m")
                pct = sc.get("pct_ebitda")
                with col:
                    st.markdown(f"""
                    <div class="card-box" style="border-top:4px solid {c}">
                      <div style="font-size:14px;font-weight:800;color:#102a43">{sc.get('name','—')}</div>
                      <div style="font-size:26px;font-weight:800;color:{c};margin-top:6px">€{imp if imp is not None else '—'}m</div>
                      <div style="font-size:12px;color:#627d98">{pct if pct is not None else '—'}% of EBITDA</div>
                      <div style="font-size:12px;color:#334e68;margin-top:8px;line-height:1.5">{sc.get('narrative','')}</div>
                    </div>""", unsafe_allow_html=True)
            sdf = pd.DataFrame([{
                "Scenario": sc.get("name", ""),
                "Financial Impact": f"€{sc.get('financial_impact_eur_m')}m" if sc.get("financial_impact_eur_m") is not None else "—",
                "% of EBITDA": f"{sc.get('pct_ebitda')}%" if sc.get("pct_ebitda") is not None else "—",
                "Narrative": sc.get("narrative", ""),
            } for sc in scenarios])
            st.dataframe(sdf, use_container_width=True, hide_index=True)
            basis = climate.get("scenario_basis") or {}
            if basis.get("ebitda_estimate_eur_m") is not None:
                st.caption(f"EBITDA basis: ~€{basis.get('ebitda_estimate_eur_m')}m "
                           f"({basis.get('ebitda_basis','estimated')}).")

    # ── Resilience Assessment ──
    with tab_resil:
        stranded = climate.get("stranded_asset_score")
        resil    = climate.get("resilience_score")
        caps     = climate.get("resilience_capabilities") or []
        if stranded is None and resil is None and not caps:
            st.info("🔄 Resilience scoring isn't available for this analysis. "
                    "Re-run the assessment to generate stranded-asset and resilience metrics.")
        else:
            g1, g2 = st.columns(2)
            with g1:
                if stranded is not None:
                    st.plotly_chart(gauge_chart(stranded, "Stranded Asset Risk"), use_container_width=True)
                    st.caption("Higher = greater stranded-asset exposure.")
            with g2:
                if resil is not None:
                    st.plotly_chart(resilience_gauge(resil, "Climate Resilience"), use_container_width=True)
                    st.caption("Higher = more resilient.")
            if caps:
                st.markdown("##### Resilience Capability Assessment")
                for cap in caps:
                    sc_val = int(cap.get("score", 0) or 0)
                    col_c  = "#2f855a" if sc_val >= 66 else "#d17f1e" if sc_val >= 33 else "#e24141"
                    st.markdown(f"""
                    <div style="margin-bottom:12px">
                      <div style="display:flex;justify-content:space-between;margin-bottom:4px">
                        <span style="font-size:13px;font-weight:700;color:#102a43">{cap.get('capability','')}</span>
                        <span style="font-size:12px;color:#627d98">{sc_val}/100</span>
                      </div>
                      <div style="height:8px;background:#eef3fb;border-radius:4px;overflow:hidden">
                        <div style="height:100%;width:{sc_val}%;background:{col_c};border-radius:4px"></div>
                      </div>
                    </div>""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# Page: ESG Engagement
# ──────────────────────────────────────────────────────────────────────────────
def page_engagement():
    r      = st.session_state.result or {}
    co     = r.get("company", {})
    risks  = r.get("risks", [])
    topics = r.get("material_topics", [])

    st.markdown('<div class="page-title">💬 ESG Engagement</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Transform assessment findings into engagement letters, questions, and action plans.</div>', unsafe_allow_html=True)

    tab_qns, tab_letter, tab_plan, tab_docs, tab_deck = st.tabs(
        ["❓ Engagement Questions", "📝 Engagement Letter", "📋 Action Plan",
         "📚 Document Library", "🎨 Board Deck"])

    with tab_qns:
        st.markdown("#### Priority Engagement Questions")
        st.caption("Generated from top risks and material topics. Tailor to your investor / lender engagement strategy.")

        top_risks  = sorted(risks, key=lambda x: -(x.get("score") or 0))[:5]
        top_topics = sorted(topics, key=lambda x: -(x.get("impact_score",0)+x.get("financial_score",0)))[:3]

        questions = []
        for risk in top_risks:
            questions.append({
                "theme": risk.get("name",""),
                "framework": risk.get("framework",""),
                "question": f"What specific actions has {co.get('name','the company')} taken to manage {risk.get('name','')} risk, "
                            f"and what quantitative targets or KPIs are in place to track progress?",
                "follow_up": f"What is your current exposure quantification under {risk.get('framework','ESRS')} and how does this compare to your 2030 targets?",
            })
        for topic in top_topics:
            questions.append({
                "theme": topic.get("topic",""),
                "framework": ", ".join(topic.get("frameworks",[]) or []),
                "question": f"Can you provide evidence of how {topic.get('topic','')} is managed as a material topic "
                            f"under your double materiality assessment?",
                "follow_up": f"What are your most recent performance metrics for {topic.get('topic','')} and how do these benchmark against sector peers?",
            })

        for i, q in enumerate(questions, 1):
            with st.expander(f"Q{i}: {q['theme']}  ·  {q['framework']}"):
                st.markdown(f"**Primary question:**")
                st.markdown(f"> {q['question']}")
                st.markdown(f"**Follow-up:**")
                st.markdown(f"> {q['follow_up']}")

    with tab_letter:
        st.markdown("#### Engagement Letter")
        company = co.get("name","[Company Name]")
        sector  = co.get("sector", "the sector")
        scores  = r.get("esg_scores", {})
        overall = scores.get("overall","—")
        country = co.get("country","—")

        letter_text = f"""**[Your Organisation Name]**
[Address]
[Date]

**ESG Engagement Notice — {company}**

Dear Management,

We are writing as [investors/lenders/creditors] in {company} to initiate a structured ESG engagement dialogue in connection with our ongoing due diligence review under [SFDR Article 8 / TCFD / PRI Responsible Investment Policy].

Our preliminary ESG risk assessment of {company} has identified an overall ESG Risk Score of **{overall}/100** across Environmental, Social, and Governance pillars. The assessment was conducted against ESRS, ISSB/IFRS S1-S2, GRI, and SASB sector standards applicable to {sector} operations in {country}.

**Priority Engagement Themes:**
{chr(10).join(f"- {r.get('name','')}: {r.get('severity','Medium')} risk — {r.get('detail','')[:80]}..." for r in sorted(risks, key=lambda x: -(x.get("score") or 0))[:3])}

We respectfully request the following disclosures within [30/60] days:

1. Quantitative GHG emissions data (Scope 1, 2, and 3) with year-on-year trend
2. Written description of board oversight and management processes for ESG risks
3. Confirmation of alignment with or plans to align to ESRS / IFRS S1-S2
4. Specific targets and timelines for the material topics identified above

We are committed to constructive dialogue and recognise the positive steps {company} has already taken. We would welcome a call or meeting at your earliest convenience.

Yours sincerely,
[Portfolio Manager / ESG Lead]
[Organisation]
"""
        st.text_area("Engagement Letter Draft", value=letter_text, height=500, key="engagement_letter")
        if st.button("📋 Copy to Clipboard"):
            st.code(letter_text, language=None)

    with tab_plan:
        st.markdown("#### Recommended Action Plan")
        st.caption("Priority actions ranked by risk severity. Use for stewardship and engagement tracking.")

        actions = []
        for risk in sorted(risks, key=lambda x: -(x.get("score") or 0)):
            if risk.get("recommendation"):
                actions.append({
                    "Priority": risk.get("severity","Medium"),
                    "Risk Topic": risk.get("name",""),
                    "Category": risk.get("category","").title(),
                    "Action": risk.get("recommendation",""),
                    "Framework": risk.get("framework",""),
                    "Score": risk.get("score",0),
                })
        if actions:
            df = pd.DataFrame(actions)
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.info("No specific recommendations available from this assessment.")

    # ── Document Library ────────────────────────────────────────────────────
    with tab_docs:
        import io, zipfile

        st.markdown("#### Document Library")
        st.caption("Ready-to-share engagement collateral — one-pagers, decks, policy templates, and data collection templates.")

        hdr_l, hdr_r = st.columns([3, 1])
        with hdr_r:
            if st.button("🔄 Refresh", key="doc_lib_refresh", use_container_width=True):
                st.session_state.pop("doc_manifest", None)
                st.session_state.pop("doc_bytes", None)

        # Fetch + cache the manifest for this session (refetch on Refresh)
        if "doc_manifest" not in st.session_state:
            try:
                st.session_state.doc_manifest = api_manifest()
            except Exception as e:
                st.error(f"Could not load the document library — is the backend running? ({e})")
                st.session_state.doc_manifest = None

        manifest = st.session_state.get("doc_manifest")
        if manifest:
            files = manifest.get("files", []) or []

            # Bucket filenames into the 4 categories by extension
            buckets = {cat: [] for cat in DOC_CATEGORY_ORDER}
            for fname in files:
                ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
                cat_info = DOC_CATEGORY_BY_EXT.get(ext)
                if cat_info:
                    buckets.setdefault(cat_info[0], []).append(fname)

            # Cache for lazily-fetched download bytes
            st.session_state.setdefault("doc_bytes", {})

            def _pretty_label(fname):
                stem = fname.rsplit(".", 1)[0] if "." in fname else fname
                return stem.replace("_", " ").replace("-", " ").title()

            # ── Download-all as a zip ──
            total = sum(len(v) for v in buckets.values())
            with hdr_l:
                st.markdown(f'<span class="badge-gray">{total} documents · {len([c for c in buckets if buckets[c]])} categories</span>',
                            unsafe_allow_html=True)

            if total:
                if st.button("⬇ Download All (.zip)", key="doc_lib_zip"):
                    try:
                        buf = io.BytesIO()
                        with st.spinner("Packaging all documents…"):
                            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                                prog = st.progress(0.0)
                                for idx, fname in enumerate(files, 1):
                                    try:
                                        zf.writestr(fname, api_engagement_file(fname))
                                    except Exception:
                                        pass  # skip individual failures, keep going
                                    prog.progress(idx / max(len(files), 1))
                                prog.empty()
                        st.session_state.doc_bytes["__zip__"] = buf.getvalue()
                    except Exception as e:
                        st.error(f"Could not build the zip archive: {e}")
                if st.session_state.doc_bytes.get("__zip__"):
                    st.download_button(
                        "💾 Save documents.zip",
                        data=st.session_state.doc_bytes["__zip__"],
                        file_name="ESGIntel_Engagement_Documents.zip",
                        mime="application/zip",
                        key="doc_lib_zip_dl",
                    )

            st.markdown("")

            # ── Per-category listing via nested tabs ──
            active_cats = [c for c in DOC_CATEGORY_ORDER if buckets.get(c)]
            if not active_cats:
                st.info("No documents are available in the library yet.")
            else:
                cat_tabs = st.tabs([f"{DOC_CATEGORY_BY_EXT[[k for k,v in DOC_CATEGORY_BY_EXT.items() if v[0]==c][0]][1]} {c} ({len(buckets[c])})"
                                    for c in active_cats])
                for ctab, cat in zip(cat_tabs, active_cats):
                    with ctab:
                        for fname in buckets[cat]:
                            ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
                            icon, badge = DOC_CATEGORY_BY_EXT.get(ext, ("📄", "badge-gray"))[1], \
                                          DOC_CATEGORY_BY_EXT.get(ext, ("", "", "badge-gray"))[2]
                            row_l, row_r = st.columns([3, 1])
                            with row_l:
                                st.markdown(
                                    f'<div class="card-box" style="padding:12px 16px;margin-bottom:8px">'
                                    f'<span style="font-size:18px">{icon}</span> '
                                    f'<strong style="color:#102a43">{_pretty_label(fname)}</strong> '
                                    f'<span class="{badge}" style="margin-left:8px">.{ext}</span>'
                                    f'<div style="font-size:11px;color:#627d98;margin-top:2px">{fname}</div>'
                                    f'</div>',
                                    unsafe_allow_html=True)
                            with row_r:
                                # Lazy two-step download: fetch on demand, then offer download_button
                                cached = st.session_state.doc_bytes.get(fname)
                                if cached is None:
                                    if st.button("Prepare download", key=f"prep_{fname}",
                                                 use_container_width=True):
                                        try:
                                            with st.spinner("Fetching…"):
                                                st.session_state.doc_bytes[fname] = api_engagement_file(fname)
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"Fetch failed: {e}")
                                else:
                                    st.download_button(
                                        "⬇ Download",
                                        data=cached,
                                        file_name=fname,
                                        mime=DOC_MIME_BY_EXT.get(ext, "application/octet-stream"),
                                        key=f"dl_{fname}",
                                        use_container_width=True,
                                    )

    # ── Board Deck (Canva) ──────────────────────────────────────────────────
    with tab_deck:
        st.markdown("#### Board Deck")
        st.caption("Generate a polished, editable board-ready ESG deck in Canva from this assessment.")

        st.markdown(
            '<div class="card-box" style="border-left:4px solid #7c3aed">'
            '<strong style="color:#102a43">🎨 Canva Board Deck</strong>'
            '<div style="font-size:12px;color:#627d98;margin-top:4px">'
            'Kicks off an async Canva render. Generation typically takes a minute or two — '
            'start it, then click <em>Check status</em> to poll for the finished deck.</div>'
            '</div>', unsafe_allow_html=True)
        st.markdown("")

        if st.button("🎨 Generate Board Deck", key="canva_generate", type="primary"):
            try:
                job = api_canva_request("board_deck")
                jid = (job or {}).get("job_id")
                if jid:
                    st.session_state.canva_job_id = jid
                    st.session_state.canva_job_started = time.time()
                    st.session_state.pop("canva_result", None)
                    st.success("Deck generation started.")
                else:
                    st.error("Backend accepted the request but returned no job id.")
            except Exception as e:
                st.error(f"Could not start deck generation — is the backend running? ({e})")

        job_id = st.session_state.get("canva_job_id")
        if job_id:
            started = st.session_state.get("canva_job_started", time.time())
            elapsed = time.time() - started
            result  = st.session_state.get("canva_result")

            st.markdown(f'<span class="badge-blue">Job {job_id}</span> '
                        f'<span class="badge-gray">{int(elapsed)}s elapsed</span>',
                        unsafe_allow_html=True)

            if result:
                url = result.get("url") or result.get("canva_url") or result.get("edit_url")
                if url:
                    st.success("Your board deck is ready.")
                    try:
                        st.link_button("🔗 Open deck in Canva", url, type="primary")
                    except Exception:
                        st.markdown(f'<a href="{url}" target="_blank">🔗 Open deck in Canva</a>',
                                    unsafe_allow_html=True)
                else:
                    st.info("Deck completed, but no Canva URL was returned.")
            else:
                if st.button("🔄 Check status", key="canva_check"):
                    try:
                        status = api_canva_status(job_id) or {}
                        state = (status.get("status") or status.get("state") or "").lower()
                        if state in ("complete", "completed", "done", "success", "succeeded"):
                            latest = api_canva_latest("board_deck") or {}
                            st.session_state.canva_result = latest
                            st.rerun()
                        else:
                            st.info(f"Status: {state or 'in progress'} — check again in a moment.")
                    except Exception as e:
                        st.error(f"Could not check status: {e}")

                if elapsed > 60:
                    st.warning("Still generating — Canva decks can take a minute or two. "
                               "Check back and click 'Check status' again.")
                else:
                    st.caption("Deck is generating… click 'Check status' to poll.")


# ──────────────────────────────────────────────────────────────────────────────
# Page: Report Generator
# ──────────────────────────────────────────────────────────────────────────────
def page_reports():
    r  = st.session_state.result or {}
    co = r.get("company", {})
    aid = st.session_state.analysis_id

    st.markdown('<div class="page-title">📄 Report Generator</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Generate professional ESG reports in Word, PDF, Excel, and PowerPoint formats.</div>', unsafe_allow_html=True)

    c1, c2 = st.columns([1, 2])

    with c1:
        st.markdown("#### Report Configuration")
        report_type = st.radio("Report Type", [
            "Full ESG Due Diligence Report",
            "Executive Summary (2 pages)",
            "Climate Risk Report (TCFD)",
            "ESG Engagement Letter",
            "Risk Register Export",
        ], key="report_type")

        st.markdown("#### Sections to Include")
        sections = {
            "Executive Summary & Analyst Narrative": True,
            "Company Profile & Sector Context": True,
            "Materiality Assessment": True,
            "Environmental Pillar Analysis": True,
            "Social Pillar Analysis": True,
            "Governance Pillar Analysis": True,
            "Climate Risk Assessment (TCFD)": True,
            "Full ESG Risk Register": True,
            "Benchmarking Analysis": True,
            "Recommended Actions": True,
            "Engagement Questions": False,
            "Scenario Analysis Financials": False,
            "Appendix: Data Sources": False,
            "Appendix: Methodology Notes": False,
        }
        selected_sections = {}
        for sec, default in sections.items():
            selected_sections[sec] = st.checkbox(sec, value=default, key=f"sec_{sec}")

        st.markdown("#### Output Format")
        fmt = st.radio("Format", ["PDF", "Word (.docx)", "Excel (.xlsx) — Risk Register", "PowerPoint (.pptx) — via Board Deck", "JSON"],
                       key="report_fmt", horizontal=True)

        st.markdown("#### Target Audience")
        audience = st.selectbox(
            "Who is this report for?",
            ["Board / Executive Committee", "Investors / LPs", "Regulators / Compliance",
             "Internal Risk Team", "Public / External Stakeholders"],
            key="report_audience",
            help="Tailors the framing of the executive narrative — currently used as report metadata; deeper per-audience rewriting is a future enhancement.",
        )

    with c2:
        st.markdown("#### Report Preview")
        company = co.get("name","[Company]")
        scores  = r.get("esg_scores",{})
        st.markdown(f"""
        <div class="card-box">
          <div style="font-size:18px;font-weight:800;color:#102a43">ESG Due Diligence Report</div>
          <div style="font-size:13px;color:#627d98;margin-top:4px">{company} · {co.get("sector","—")} · {co.get("country","—")}</div>
          <hr style="border-color:rgba(145,158,171,0.14)">
          <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-top:12px">
            <div style="text-align:center"><div style="font-size:11px;color:#627d98;text-transform:uppercase">Overall</div><div style="font-size:24px;font-weight:800;color:{score_color(scores.get('overall',0))}">{scores.get('overall','—')}</div></div>
            <div style="text-align:center"><div style="font-size:11px;color:#627d98;text-transform:uppercase">Environmental</div><div style="font-size:24px;font-weight:800;color:{score_color(scores.get('environmental',0))}">{scores.get('environmental','—')}</div></div>
            <div style="text-align:center"><div style="font-size:11px;color:#627d98;text-transform:uppercase">Social</div><div style="font-size:24px;font-weight:800;color:{score_color(scores.get('social',0))}">{scores.get('social','—')}</div></div>
          </div>
          <div style="margin-top:16px;font-size:13px;color:#334e68">
            <strong>Sections:</strong> {", ".join(k for k,v in selected_sections.items() if v)[:120]}…
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("#### Download Report")
        if not aid:
            st.warning("No active analysis. Run an assessment first.")
            return

        if not backend_ok():
            st.error("Backend is offline. Start the FastAPI server first.")
            return

        # Maps the radio label -> backend `type` key (backend/documents/company_data.py REPORT_TYPES)
        REPORT_TYPE_KEYS = {
            "Full ESG Due Diligence Report": "full",
            "Executive Summary (2 pages)":   "exec",
            "Climate Risk Report (TCFD)":    "climate",
            "ESG Engagement Letter":         "engagement",
            "Risk Register Export":          "riskregister",
        }
        report_type_key = REPORT_TYPE_KEYS.get(report_type, "full")

        # NOTE: GET /api/documents/report only supports fmt="docx"|"pdf" and always
        # serves the backend's *latest* stored assessment (it has no analysis_id
        # param) — if you're viewing an older analysis from "Past Analyses", the
        # downloaded report reflects the most recent run, not necessarily this one.
        fmt_endpoints = {
            "Word (.docx)": ("docx", "docx", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
            "PDF":          ("pdf",  "pdf",  "application/pdf"),
        }

        if st.button(f"🔄 Generate {fmt} Report", type="primary", use_container_width=True):
            if fmt == "JSON":
                import json
                json_bytes = json.dumps(r, indent=2, ensure_ascii=False).encode("utf-8")
                st.download_button("⬇ Download JSON", data=json_bytes,
                                   file_name=f"ESGIntel_{company.replace(' ','_')}.json",
                                   mime="application/json")
            elif fmt.startswith("Excel"):
                # No backend xlsx report endpoint exists — reuse the local
                # openpyxl risk-register export built for the Risk Register page.
                risks = r.get("risks", [])
                if not risks:
                    st.warning("No risk data available in this assessment to export.")
                else:
                    try:
                        xls = _risk_register_xlsx(risks, company)
                        cname = company.replace(" ","_").replace("/","_")
                        st.download_button(
                            "⬇ Download Risk Register (.xlsx)",
                            data=xls,
                            file_name=f"ESGIntel_{cname}_RiskRegister.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                        st.caption("Excel export always contains the full risk register, independent of the report type selected above.")
                    except Exception as e:
                        st.error(f"Excel export failed: {e}")
            elif fmt.startswith("PowerPoint"):
                st.info("PowerPoint decks are generated via the Canva integration — go to **Engagement → Generate Board Deck** to create one from this analysis.")
            else:
                endpoint_info = fmt_endpoints.get(fmt)
                if not endpoint_info:
                    st.error(f"Unsupported format: {fmt}")
                elif not aid:
                    st.warning("No active analysis. Run an assessment first.")
                else:
                    backend_fmt, ext, mime = endpoint_info
                    with st.spinner(f"Generating {fmt} report…"):
                        try:
                            resp = requests.get(
                                f"{API_BASE}/api/documents/report",
                                params={"type": report_type_key, "fmt": backend_fmt},
                                timeout=60,
                            )
                            if resp.status_code == 200:
                                cname = company.replace(" ","_").replace("/","_")
                                st.download_button(
                                    f"⬇ Download {fmt}",
                                    data=resp.content,
                                    file_name=f"ESGIntel_{cname}_{report_type_key}.{ext}",
                                    mime=mime,
                                )
                            else:
                                st.error(f"Report generation failed (HTTP {resp.status_code}): {resp.text[:200]}")
                        except requests.exceptions.ConnectionError:
                            st.error("Could not connect to backend. Is the FastAPI server running?")
                        except Exception as e:
                            st.error(f"Error: {e}")

        # Quick JSON download always available
        import json as json_mod
        if r:
            json_bytes = json_mod.dumps(r, indent=2, ensure_ascii=False).encode("utf-8")
            cname = company.replace(" ","_").replace("/","_")
            st.download_button(
                "⬇ Download Raw JSON (always available)",
                data=json_bytes,
                file_name=f"ESGIntel_{cname}_raw.json",
                mime="application/json",
                key="json_dl_2",
            )


# ──────────────────────────────────────────────────────────────────────────────
# Account & Admin — authenticated API helpers
# ──────────────────────────────────────────────────────────────────────────────
SUPPORT_EMAIL = "support@esgintel.io"


def _auth_headers():
    """Authorization header for the current session's JWT (empty if none)."""
    tok = st.session_state.get("auth_token")
    return {"Authorization": f"Bearer {tok}"} if tok else {}


class AuthExpired(Exception):
    """Raised on a 401 so callers can prompt a re-login instead of crashing."""


class AdminForbidden(Exception):
    """Raised on a 403 from an admin endpoint (caller isn't an admin)."""


def _guard(resp):
    """Turn 401/403 into typed exceptions; leave other statuses to the caller."""
    if resp.status_code == 401:
        raise AuthExpired("Your session has expired — please log out and sign in again.")
    if resp.status_code == 403:
        raise AdminForbidden("Admin access required.")
    return resp


def _api_json(method, path, *, json_body=None, params=None, timeout=20):
    """Authenticated JSON request → parsed body. Raises AuthExpired/AdminForbidden
    on 401/403, RuntimeError with the backend detail on other 4xx/5xx."""
    resp = requests.request(method, f"{API_BASE}{path}", headers=_auth_headers(),
                            json=json_body, params=params, timeout=timeout)
    _guard(resp)
    if resp.status_code >= 400:
        raise RuntimeError(_auth_error(resp))
    if resp.status_code == 204 or not (resp.content or b"").strip():
        return {}
    try:
        return resp.json()
    except Exception:
        return {}


def _render_auth_expired():
    """Shared UI shown when the backend rejects our token (401)."""
    st.error("🔒 Your session has expired. Please sign in again to continue.")
    if st.button("↩ Log out and sign in", key="auth_expired_logout", type="primary"):
        st.session_state.auth_token = None
        st.session_state.user = None
        st.rerun()


def _is_admin_user(user: dict) -> bool:
    user = user or {}
    return bool(user.get("isAdmin") or user.get("is_admin")
                or (user.get("role", "") or "").lower() == "admin")


# ──────────────────────────────────────────────────────────────────────────────
# Page: Account settings
# ──────────────────────────────────────────────────────────────────────────────
def page_account():
    st.markdown('<div class="page-title">👤 Account</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Manage your profile, notifications, API keys and team.</div>',
                unsafe_allow_html=True)

    if not backend_ok():
        st.error("Backend is offline. Start the FastAPI server at :8000 to manage account settings.")
        return

    # Load the profile once so Profile + Notifications tabs share display name / company.
    profile, profile_err = {}, None
    try:
        profile = _api_json("GET", "/api/account/profile") or {}
    except AuthExpired:
        _render_auth_expired()
        return
    except Exception as e:
        profile_err = str(e)

    disp  = profile.get("displayName") or profile.get("display_name") or ""
    comp  = profile.get("company") or ""
    prefs = profile.get("notificationPrefs") or profile.get("notification_prefs") or {}

    tabs = st.tabs(["My Profile", "Notifications", "API Keys", "Team", "Support"])

    # ── My Profile ──
    with tabs[0]:
        st.markdown("#### Personal information")
        if profile_err:
            st.warning(f"Could not load your profile ({profile_err}). You can still save changes below.")
        with st.form("account_profile_form"):
            new_disp = st.text_input("Display name", value=disp, placeholder="Jane Analyst")
            new_comp = st.text_input("Company / Fund", value=comp, placeholder="Apex Capital Management")
            email    = (st.session_state.get("user") or {}).get("email", "")
            if email:
                st.text_input("Email address", value=email, disabled=True,
                              help="Contact support to change the email on your account.")
            saved = st.form_submit_button("Save changes", type="primary", use_container_width=True)
        if saved:
            try:
                _api_json("PUT", "/api/account/profile", json_body={
                    "displayName": new_disp, "company": new_comp,
                    "notificationPrefs": prefs,
                })
                # keep the cached user (topbar) in sync
                u = st.session_state.get("user") or {}
                u["displayName"] = new_disp
                u["company"] = new_comp
                st.session_state.user = u
                st.success("Profile saved.")
            except AuthExpired:
                _render_auth_expired()
            except Exception as e:
                st.error(f"Could not save profile: {e}")

    # ── Notifications ──
    with tabs[1]:
        st.markdown("#### Email notifications")
        st.caption("Choose which emails ESGIntel sends you. Saved to your profile.")
        if profile_err:
            st.warning(f"Could not load your notification preferences ({profile_err}).")
        NOTIF_FIELDS = [
            ("assessment_complete", "Email me when an analysis completes", True),
            ("weekly_digest",       "Weekly ESG digest (every Monday)",    False),
            ("regulatory_updates",  "Regulatory updates (ESRS, CBAM, EU ETS)", True),
            ("team_activity",       "Team activity (members run or share reports)", True),
            ("product_updates",     "Product updates & release notes",     True),
        ]
        with st.form("account_notif_form"):
            new_prefs = {}
            for key, label, default in NOTIF_FIELDS:
                current = prefs.get(key, default) if isinstance(prefs, dict) else default
                new_prefs[key] = st.toggle(label, value=bool(current), key=f"notif_{key}")
            saved_n = st.form_submit_button("Save preferences", type="primary", use_container_width=True)
        if saved_n:
            try:
                merged = dict(prefs) if isinstance(prefs, dict) else {}
                merged.update(new_prefs)
                _api_json("PUT", "/api/account/profile", json_body={
                    "displayName": disp, "company": comp, "notificationPrefs": merged,
                })
                st.success("Notification preferences saved.")
            except AuthExpired:
                _render_auth_expired()
            except Exception as e:
                st.error(f"Could not save preferences: {e}")

    # ── API Keys ──
    with tabs[2]:
        st.markdown("#### API keys")
        st.caption("Use these keys to access the ESGIntel API from your own systems. "
                   "A key's secret is shown only once, at creation.")

        # Show a freshly-created key exactly once (persisted across the form rerun).
        new_key = st.session_state.get("_new_api_key")
        if new_key:
            st.success(f"New key **{new_key.get('label','')}** created — copy it now, "
                       "it will not be shown again.")
            st.code(new_key.get("key", ""), language="text")
            if st.button("I've copied it — dismiss", key="dismiss_new_key"):
                st.session_state.pop("_new_api_key", None)
                st.rerun()

        with st.form("account_newkey_form"):
            kc1, kc2 = st.columns([3, 1])
            with kc1:
                key_label = st.text_input("Key label", placeholder="e.g. Production integration",
                                          label_visibility="collapsed")
            with kc2:
                gen = st.form_submit_button("＋ Generate key", type="primary", use_container_width=True)
        if gen:
            if not (key_label or "").strip():
                st.error("Enter a label for the new key.")
            else:
                try:
                    data = _api_json("POST", "/api/account/api-keys",
                                     json_body={"label": key_label.strip()})
                    plaintext = (data.get("key") or data.get("plaintext")
                                 or data.get("api_key") or data.get("secret") or "")
                    st.session_state["_new_api_key"] = {
                        "label": data.get("label", key_label.strip()), "key": plaintext,
                    }
                    st.rerun()
                except AuthExpired:
                    _render_auth_expired()
                except Exception as e:
                    st.error(f"Could not create key: {e}")

        st.markdown("##### Existing keys")
        try:
            keys = _api_json("GET", "/api/account/api-keys") or []
            if isinstance(keys, dict):
                keys = keys.get("keys") or keys.get("items") or []
            if not keys:
                st.caption("No API keys yet.")
            for k in keys:
                kid = k.get("id") or k.get("key_id")
                c1, c2, c3, c4 = st.columns([2.2, 1.6, 1.6, 1])
                c1.markdown(f"**{k.get('label','(unnamed)')}**")
                c2.caption(f"Created {str(k.get('created_at','—'))[:10]}")
                last = k.get("last_used_at")
                c3.caption(f"Last used {str(last)[:10] if last else 'never'}")
                if c4.button("Revoke", key=f"revoke_{kid}"):
                    try:
                        _api_json("DELETE", f"/api/account/api-keys/{kid}")
                        st.toast("Key revoked.")
                        st.rerun()
                    except AuthExpired:
                        _render_auth_expired()
                    except Exception as e:
                        st.error(f"Could not revoke key: {e}")
        except AuthExpired:
            _render_auth_expired()
        except Exception as e:
            st.error(f"Could not load API keys: {e}")

    # ── Team ──
    with tabs[3]:
        st.markdown("#### Team & permissions")
        st.caption("Invite teammates and manage pending invitations.")
        TEAM_ROLES = ["admin", "manager", "analyst", "viewer"]
        with st.form("account_invite_form"):
            ic1, ic2, ic3 = st.columns([2.4, 1.4, 1])
            with ic1:
                invite_email = st.text_input("Email", placeholder="colleague@firm.com",
                                             label_visibility="collapsed")
            with ic2:
                invite_role = st.selectbox("Role", TEAM_ROLES,
                                           format_func=str.title, label_visibility="collapsed")
            with ic3:
                sent = st.form_submit_button("Invite", type="primary", use_container_width=True)
        if sent:
            if not (invite_email or "").strip() or "@" not in invite_email:
                st.error("Enter a valid email address.")
            else:
                try:
                    _api_json("POST", "/api/account/team/invite",
                              json_body={"email": invite_email.strip(), "role": invite_role})
                    st.success(f"Invitation sent to {invite_email.strip()}.")
                    st.rerun()
                except AuthExpired:
                    _render_auth_expired()
                except Exception as e:
                    st.error(f"Could not send invite: {e}")

        st.markdown("##### Current team & invitations")
        try:
            team = _api_json("GET", "/api/account/team") or []
            if isinstance(team, dict):
                team = team.get("members") or team.get("invites") or team.get("items") or []
            if not team:
                st.caption("No team members or pending invitations yet.")
            for m in team:
                mid = m.get("id") or m.get("invite_id")
                c1, c2, c3, c4 = st.columns([2.4, 1.3, 1.3, 1])
                c1.markdown(f"**{m.get('email','—')}**")
                c2.markdown(f'<span class="badge-blue">{(m.get("role","") or "—").title()}</span>',
                            unsafe_allow_html=True)
                status = (m.get("status") or "pending").title()
                badge = "badge-green" if status.lower() in ("active", "accepted") else "badge-amber"
                c3.markdown(f'<span class="{badge}">{status}</span>', unsafe_allow_html=True)
                # Only pending invites can be cancelled
                if mid and status.lower() not in ("active", "accepted"):
                    if c4.button("Cancel", key=f"cancel_invite_{mid}"):
                        try:
                            _api_json("DELETE", f"/api/account/team/invite/{mid}")
                            st.toast("Invitation cancelled.")
                            st.rerun()
                        except AuthExpired:
                            _render_auth_expired()
                        except Exception as e:
                            st.error(f"Could not cancel invite: {e}")
        except AuthExpired:
            _render_auth_expired()
        except Exception as e:
            st.error(f"Could not load team: {e}")

    # ── Support ──
    with tabs[4]:
        st.markdown("#### Support & help")
        st.markdown(f"""
        <div class="card-box">
          <div style="font-size:14px;font-weight:800;color:#102a43">Need a hand?</div>
          <div style="font-size:13px;color:#334e68;margin-top:8px;line-height:1.7">
            Our team typically responds within one business day.<br>
            📧 Email us at <a href="mailto:{SUPPORT_EMAIL}">{SUPPORT_EMAIL}</a><br>
            📖 Browse the documentation for API references and methodology notes.<br>
            💬 In-app chat is available on Professional and Enterprise plans.
          </div>
        </div>
        """, unsafe_allow_html=True)
        st.link_button("Email support", f"mailto:{SUPPORT_EMAIL}", use_container_width=False)


# ──────────────────────────────────────────────────────────────────────────────
# Page: Admin panel (admin-only)
# ──────────────────────────────────────────────────────────────────────────────
def page_admin():
    st.markdown('<div class="page-title">⚙️ Admin Panel</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Platform administration — users, plans and stats.</div>',
                unsafe_allow_html=True)

    # ── Gate: only render for admins; never touch admin endpoints otherwise ──
    if not _is_admin_user(st.session_state.get("user")):
        st.error("Admin access required.")
        st.caption("Your account doesn't have administrator privileges.")
        return

    if not backend_ok():
        st.error("Backend is offline. Start the FastAPI server at :8000 to use the admin panel.")
        return

    # ── Stats strip ──
    try:
        stats = _api_json("GET", "/api/admin/stats") or {}
    except AuthExpired:
        _render_auth_expired()
        return
    except AdminForbidden:
        st.error("Admin access required.")
        return
    except Exception as e:
        stats = {}
        st.warning(f"Could not load platform stats: {e}")

    def _stat(*keys, default="—"):
        for k in keys:
            if stats.get(k) is not None:
                return stats.get(k)
        return default

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Users",     _stat("total_users", "totalUsers", "users"))
    m2.metric("Total Analyses",  _stat("total_analyses", "totalAnalyses", "analyses"))
    m3.metric("Active (7d)",     _stat("active_7d", "new_this_week", "active7d", "newThisWeek"))
    m4.metric("Active (30d)",    _stat("active_30d", "active30d", "pro_users", "proUsers"))

    st.markdown("---")

    # ── Users table + search ──
    hc1, hc2 = st.columns([3, 1])
    with hc1:
        st.markdown("#### All users")
    search = st.text_input("Search users", key="admin_search",
                           placeholder="Search name, email, company…", label_visibility="collapsed")

    try:
        users = _api_json("GET", "/api/admin/users",
                          params={"search": search} if search else None) or []
        if isinstance(users, dict):
            users = users.get("users") or users.get("items") or []
    except AuthExpired:
        _render_auth_expired()
        return
    except AdminForbidden:
        st.error("Admin access required.")
        return
    except Exception as e:
        users = []
        st.error(f"Could not load users: {e}")

    PLAN_OPTIONS = ["free", "starter", "professional", "enterprise"]

    if not users:
        st.caption("No users found.")
    else:
        # Build a display frame; the Plan column is editable inline via data_editor.
        def _uid(u):
            return u.get("id") or u.get("uid") or u.get("user_id")

        rows, id_by_index = [], []
        for u in users:
            id_by_index.append(_uid(u))
            plan = (u.get("plan") or "free").lower()
            rows.append({
                "Name":    u.get("displayName") or u.get("display_name") or "—",
                "Email":   u.get("email", "—"),
                "Company": u.get("company", "—"),
                "Role":    (u.get("role", "") or "—").title(),
                "Plan":    plan if plan in PLAN_OPTIONS else "free",
                "Joined":  str(u.get("created_at") or u.get("joined") or "—")[:10],
            })
        df = pd.DataFrame(rows)
        original_plans = df["Plan"].tolist()

        edited = st.data_editor(
            df, use_container_width=True, hide_index=True, key="admin_users_editor",
            disabled=["Name", "Email", "Company", "Role", "Joined"],
            column_config={
                "Plan": st.column_config.SelectboxColumn(
                    "Plan", options=PLAN_OPTIONS, required=True,
                    help="Change a plan, then click Apply plan changes."),
            },
        )

        ac1, ac2 = st.columns([1, 3])
        with ac1:
            if st.button("Apply plan changes", type="primary", use_container_width=True):
                changed = 0
                errors = []
                for idx, new_plan in enumerate(edited["Plan"].tolist()):
                    if new_plan != original_plans[idx]:
                        uid = id_by_index[idx]
                        if not uid:
                            continue
                        try:
                            _api_json("PATCH", f"/api/admin/users/{uid}/plan",
                                      json_body={"plan": new_plan})
                            changed += 1
                        except AuthExpired:
                            _render_auth_expired()
                            return
                        except Exception as e:
                            errors.append(f"{edited['Email'].iloc[idx]}: {e}")
                if changed:
                    st.success(f"Updated {changed} plan(s).")
                if errors:
                    st.error("Some updates failed:\n\n" + "\n".join(errors))
                if changed and not errors:
                    st.rerun()

    st.markdown("---")

    # ── Export CSV ──
    ec1, ec2 = st.columns([1, 3])
    with ec1:
        if st.button("⬇ Fetch users CSV", use_container_width=True):
            try:
                resp = requests.get(f"{API_BASE}/api/admin/users/export",
                                    headers=_auth_headers(), timeout=30)
                _guard(resp)
                if resp.status_code == 200:
                    st.session_state["_admin_csv"] = resp.content
                else:
                    st.error(f"Export failed (HTTP {resp.status_code}).")
            except AuthExpired:
                _render_auth_expired()
            except AdminForbidden:
                st.error("Admin access required.")
            except Exception as e:
                st.error(f"Could not export users: {e}")
    with ec2:
        if st.session_state.get("_admin_csv"):
            st.download_button("Download users.csv", data=st.session_state["_admin_csv"],
                               file_name="esgintel_users.csv", mime="text/csv")

    st.markdown("---")

    # ── Platform settings (session-local demo only) ──
    st.markdown("#### Platform settings")
    st.caption("These toggles are local to your current session only (demo) — there is no "
               "backend persistence endpoint for platform settings in this build.")
    st.session_state.setdefault("_ps_signups", True)
    st.session_state.setdefault("_ps_verify", False)
    ps1, ps2 = st.columns(2)
    with ps1:
        st.toggle("Allow new user signups", key="_ps_signups")
    with ps2:
        st.toggle("Require email verification", key="_ps_verify")


# ──────────────────────────────────────────────────────────────────────────────
# Router
# ──────────────────────────────────────────────────────────────────────────────
def main():
    # ── OAuth callback: capture ?token=… that the backend redirects back with ──
    qp = st.query_params
    if "token" in qp and not st.session_state.get("auth_token"):
        st.session_state.auth_token = qp["token"]
        try:
            st.session_state.user = api_me(qp["token"])
        except Exception:
            st.session_state.user = {}
        st.query_params.clear()  # strip the token from the URL after capturing it
        st.rerun()

    # ── Auth gate: show only the login screen until authenticated ──
    if not st.session_state.get("auth_token"):
        render_auth()
        return

    # ── Authenticated app: persistent topbar + sidebar + routed content ──
    render_topbar()
    render_sidebar()

    page = st.session_state.page

    if page == "input":
        page_input()
    elif page == "running":
        page_running()
    elif page == "dashboard":
        page_dashboard()
    elif page == "profile":
        page_profile()
    elif page == "materiality":
        page_materiality()
    elif page == "esg_risk":
        page_esg_risk()
    elif page == "risk_register":
        page_risk_register()
    elif page == "environmental":
        page_pillar("environmental")
    elif page == "social":
        page_pillar("social")
    elif page == "governance":
        page_pillar("governance")
    elif page == "climate":
        page_climate()
    elif page == "engagement":
        page_engagement()
    elif page == "reports":
        page_reports()
    elif page == "account":
        page_account()
    elif page == "admin":
        page_admin()
    else:
        st.error(f"Unknown page: {page}")
        st.session_state.page = "input"
        st.rerun()


if __name__ == "__main__":
    main()
