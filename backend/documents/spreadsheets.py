"""
ESGIntel — Spreadsheet generators (.xlsx)
=========================================
Produces the workbook deliverables surfaced in the Engagement tab:
  • esrs_kpi      → ESRS KPI Data Collection Template
  • supplier_saq  → Supplier ESG Self-Assessment (SAQ)
  • risk_register → ESG Risk Register export

Public contract:
    build_spreadsheet(key: str, out_dir: str) -> str
"""

import os
import company_data as cd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── house style ──────────────────────────────
INK    = cd.BRAND["ink"]
ACCENT = cd.BRAND["accent"]
ACC_LT = cd.BRAND["accent_lt"]
RED    = cd.BRAND["red"]
AMBER  = cd.BRAND["amber"]
GREEN  = cd.BRAND["green"]
RULE   = cd.BRAND["rule"]

THIN = Side(style="thin", color=RULE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor=ACCENT)
BAND_FILL = PatternFill("solid", fgColor=ACC_LT)
WHITE = Font(color="FFFFFF", bold=True, size=11)
H1 = Font(color=INK, bold=True, size=16)
H2 = Font(color=INK, bold=True, size=12)
SUB = Font(color="6B7280", size=10, italic=True)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def _title_block(ws, title, subtitle, last_col=4):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, f"ESGIntel  ·  {cd.COMPANY['name']}")
    c.font = Font(color=ACCENT, bold=True, size=11)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    t = ws.cell(2, 1, title); t.font = H1
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    s = ws.cell(3, 1, subtitle); s.font = SUB
    return 5  # first free row


def _header_row(ws, row, headers, widths=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row, j, h)
        c.fill = HEAD_FILL; c.font = WHITE; c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        if widths:
            ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
    ws.row_dimensions[row].height = 26


def _rag_fill(status):
    return PatternFill("solid", fgColor={"red": "FBE3E0", "amber": "FBEFD8",
                                         "green": "E1F0E7"}.get(status, "FFFFFF"))


def _sev_fill(sev):
    return {"Critical": PatternFill("solid", fgColor="FBE3E0"),
            "High": PatternFill("solid", fgColor="FBEFD8"),
            "Medium": PatternFill("solid", fgColor="FFF7E6"),
            "Low": PatternFill("solid", fgColor="E1F0E7")}.get(sev, PatternFill())


def _freeze(ws, cell):
    ws.freeze_panes = cell


# ─────────────────────────────────────────────
# 1) ESRS KPI Data Collection Template
# ─────────────────────────────────────────────
def _esrs_kpi(path):
    wb = Workbook()
    # Instructions / cover sheet
    ws0 = wb.active; ws0.title = "Read me"
    r = _title_block(ws0, "ESRS KPI Data Collection Template",
                     "Populate the highlighted cells. One tab per ESRS topic. " + cd.COMPANY["as_of"], 2)
    notes = [
        "How to use this workbook:",
        "•  Enter your reported figure in the 'FY2023 (enter)' column and your target in 'Target'.",
        "•  Baseline figures pre-filled from the ESGIntel assessment where available.",
        "•  Cells shaded green are inputs you complete; leave methodology notes in the final column.",
        "•  Units are specified per row — keep them consistent for assurance.",
        "",
        "Tabs:  E1 Climate  ·  E3 Water  ·  E5 Circular  ·  S1 Workforce  ·  G1 Governance",
    ]
    for n in notes:
        ws0.cell(r, 1, n).font = BOLD if n.endswith(":") else Font(size=10)
        r += 1
    ws0.column_dimensions["A"].width = 90

    # Build topic tabs from KPIS grouped
    groups = {}
    for k in cd.KPIS:
        groups.setdefault(k["group"], []).append(k)

    tab_names = {
        "Climate (E1)": "E1 Climate", "Water (E3)": "E3 Water",
        "Circular (E5)": "E5 Circular", "Workforce (S1)": "S1 Workforce",
        "Governance (G1)": "G1 Governance",
    }
    for group, rows in groups.items():
        ws = wb.create_sheet(tab_names.get(group, group[:28]))
        top = _title_block(ws, group, "ESRS data points — baseline vs your input vs target", 6)
        headers = ["KPI / data point", "Unit / basis", "Baseline (assessment)",
                   "FY2023 (enter)", "Target", "Methodology note"]
        _header_row(ws, top, headers, widths=[34, 18, 20, 16, 14, 30])
        rr = top + 1
        for k in rows:
            ws.cell(rr, 1, k["metric"]).border = BORDER
            ws.cell(rr, 2, "see value").border = BORDER
            base = ws.cell(rr, 3, k["value"]); base.border = BORDER
            inp = ws.cell(rr, 4, ""); inp.border = BORDER; inp.fill = BAND_FILL
            tgt = ws.cell(rr, 5, ""); tgt.border = BORDER; tgt.fill = BAND_FILL
            note = ws.cell(rr, 6, ""); note.border = BORDER; note.alignment = WRAP
            if k["pctl"] is not None:
                ws.cell(rr, 2, f"peer P{k['pctl']}")
            rr += 1
        _freeze(ws, "A" + str(top + 1))
    wb.save(path)
    return path


# ─────────────────────────────────────────────
# 2) Supplier ESG Self-Assessment (SAQ)
# ─────────────────────────────────────────────
def _supplier_saq(path):
    wb = Workbook()
    ws = wb.active; ws.title = "Supplier SAQ"
    top = _title_block(ws, "Supplier ESG Self-Assessment Questionnaire (SAQ)",
                       "CSDDD-ready · Tier-1 supplier due diligence · " + cd.COMPANY["as_of"], 4)
    # Company info block
    ws.cell(top, 1, "Section 1 — Supplier information").font = H2
    top += 1
    info = ["Company name", "Country", "Primary product / service", "Employees",
            "Annual revenue (EUR)", "Primary contact / role"]
    for label in info:
        ws.cell(top, 1, label).font = BOLD
        ws.merge_cells(start_row=top, start_column=2, end_row=top, end_column=4)
        ws.cell(top, 2, "").fill = BAND_FILL
        for col in (1, 2):
            ws.cell(top, col).border = BORDER
        top += 1
    top += 1

    sections = [
        ("Section 2 — Environmental", [
            "Do you have a documented Environmental Policy?",
            "Are you ISO 14001 certified?",
            "Do you measure Scope 1 & 2 GHG emissions?",
            "Do you have a net-zero / decarbonisation target?",
            "Do any operations sit in water-stressed areas?",
        ]),
        ("Section 3 — Labour & Human Rights", [
            "Is child labour prohibited and verified?",
            "Is forced / bonded labour prohibited and verified?",
            "Is freedom of association respected?",
            "Do you have a Human Rights Policy (UNGP-aligned)?",
            "Do you pay at least a living wage?",
            "Do you operate a worker grievance mechanism?",
        ]),
        ("Section 4 — Governance & Ethics", [
            "Do you have a Code of Conduct?",
            "Do you have an anti-bribery & corruption policy?",
            "Do you operate a whistleblower mechanism?",
            "Any material ESG violations in the past 3 years?",
        ]),
    ]
    for title, qs in sections:
        ws.cell(top, 1, title).font = H2
        top += 1
        _header_row(ws, top, ["#", "Question", "Response (Yes/No)", "Evidence / comment"],
                    widths=[5, 60, 18, 34])
        top += 1
        for i, q in enumerate(qs, 1):
            ws.cell(top, 1, i).border = BORDER
            qc = ws.cell(top, 2, q); qc.border = BORDER; qc.alignment = WRAP
            rc = ws.cell(top, 3, ""); rc.border = BORDER; rc.fill = BAND_FILL; rc.alignment = CENTER
            ec = ws.cell(top, 4, ""); ec.border = BORDER; ec.fill = BAND_FILL; ec.alignment = WRAP
            top += 1
        top += 1

    ws.cell(top, 1, "Declaration").font = H2
    top += 1
    for label in ["Authorised name", "Title", "Date", "Signature"]:
        ws.cell(top, 1, label).font = BOLD
        ws.merge_cells(start_row=top, start_column=2, end_row=top, end_column=4)
        ws.cell(top, 2, "").fill = BAND_FILL
        ws.cell(top, 1).border = BORDER; ws.cell(top, 2).border = BORDER
        top += 1
    _freeze(ws, "A6")
    wb.save(path)
    return path


# ─────────────────────────────────────────────
# 3) ESG Risk Register export
# ─────────────────────────────────────────────
def _normalise_risk(risk: dict) -> dict:
    """
    Normalise a risk dict from either the legacy demo schema or the AI schema.

    Legacy demo fields:  id, pillar (E/S/G), title, score, financial, horizon, action
    AI schema fields:    name, category (environmental/social/governance/climate),
                         detail, recommendation, framework, severity, evidence, kpis

    Returns a unified dict with: id, pillar, title, score, severity,
                                  financial, horizon, action
    """
    score = risk.get("score", 0)

    # Pillar: old schema uses single letter; new schema uses full word
    raw_pillar = risk.get("pillar") or risk.get("category", "environmental")
    cat_map = {
        "environmental": "E", "climate": "E",
        "social": "S",
        "governance": "G",
    }
    pillar = cat_map.get(raw_pillar.lower(), raw_pillar.upper()[:1]) if raw_pillar else "E"
    pillar_label = {"E": "Env", "S": "Social", "G": "Gov"}.get(pillar, pillar)

    # Title / name
    title = risk.get("title") or risk.get("name") or "Unknown Risk"

    # Severity
    sev = risk.get("severity") or cd.severity(score)

    # Financial exposure: old schema has explicit field; AI schema embeds in detail
    financial = risk.get("financial") or risk.get("financial_exposure") or "See detail"

    # Horizon: old schema explicit; AI schema may embed in detail
    horizon = risk.get("horizon") or "—"

    # Action / recommendation
    action = risk.get("action") or risk.get("recommendation") or "—"

    # ID
    risk_id = risk.get("id") or f"{pillar}-{str(score).zfill(2)}"

    return {
        "id": risk_id,
        "pillar": pillar_label,
        "title": title,
        "score": score,
        "severity": sev,
        "financial": financial,
        "horizon": horizon,
        "action": action,
    }


def _risk_register(path):
    wb = Workbook()
    ws = wb.active; ws.title = "Risk Register"
    overall = cd.SCORES.get("overall", "—")
    percentile = cd.SCORES.get("overall_percentile", "")
    percentile_str = f" (P{percentile})" if percentile else ""
    as_of = cd.COMPANY.get("as_of", "")
    top = _title_block(ws, "ESG Risk Register",
                       f"Overall ESG risk {overall}/100{percentile_str} · "
                       f"{len(cd.RISKS)} active risks" + (f" · {as_of}" if as_of else ""), 8)
    headers = ["ID", "Pillar", "Risk", "Score", "Severity", "Financial exposure",
               "Horizon", "Recommended action"]
    _header_row(ws, top, headers, widths=[9, 7, 30, 8, 11, 24, 22, 40])
    rr = top + 1

    normalised = sorted(
        [_normalise_risk(r) for r in cd.RISKS],
        key=lambda x: -(x["score"] or 0),
    )

    for risk in normalised:
        sev = risk["severity"]
        vals = [risk["id"], risk["pillar"], risk["title"], risk["score"],
                sev, risk["financial"], risk["horizon"], risk["action"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(rr, j, v); c.border = BORDER
            c.alignment = WRAP if j in (3, 6, 7, 8) else (CENTER if j in (2, 4, 5) else Alignment(vertical="top"))
            if j == 5:
                c.fill = _sev_fill(sev); c.font = BOLD
            if j == 4:
                c.font = BOLD
        rr += 1

    # Severity summary block
    rr += 1
    ws.cell(rr, 1, "Severity distribution").font = H2
    rr += 1
    from collections import Counter
    counts = Counter(r["severity"] for r in normalised)
    _header_row(ws, rr, ["Severity", "Count"], widths=[14, 8])
    rr += 1
    for sev in ["Critical", "High", "Medium", "Low"]:
        a = ws.cell(rr, 1, sev); a.border = BORDER; a.fill = _sev_fill(sev); a.font = BOLD
        b = ws.cell(rr, 2, counts.get(sev, 0)); b.border = BORDER; b.alignment = CENTER
        rr += 1
    _freeze(ws, "A" + str(top + 1))
    wb.save(path)
    return path


# ─────────────────────────────────────────────
# Public dispatcher
# ─────────────────────────────────────────────
_BUILDERS = {
    "esrs_kpi":      ("ESRS_KPI_Data_Collection_Template", _esrs_kpi),
    "supplier_saq":  ("Supplier_ESG_Self_Assessment_SAQ", _supplier_saq),
    "risk_register": ("ESG_Risk_Register", _risk_register),
}


def build_spreadsheet(key: str, out_dir: str) -> str:
    if key not in _BUILDERS:
        raise ValueError(f"Unknown spreadsheet key: {key}")
    name, fn = _BUILDERS[key]
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"{cd.COMPANY['short']}_{name}.xlsx")
    return fn(path)
